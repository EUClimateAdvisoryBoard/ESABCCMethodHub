#!/usr/bin/env python3
"""Generate src/data/sector-targets.generated.ts from the policy-targets
Masterfile (public/data/eu-policy-targets-master.xlsx, tab "Clean targets").

The Clean targets tab is the analysis dataset of the register: objectives and
targets still in effect (timeline beyond 2026) that are not duplicates or
cross-references, after the August 2026 (v5) correction pass. Each row carries
a human-authored layman "Short version" and, where the v5 pass changed
anything, a "Revision note".

This script adds the machine side: a stable id, a sector list read off the
eight sector columns (no sector = cross-cutting), and a temporal encoding for
the timeline figures (kind: by / from / range / stepped / periodic, plus start
and end achievement years). The temporal encodings were authored together with
the short versions and are keyed by (act identifier, target #) below; a row
without an entry falls back to a regex year scan of its timeline text.

Run: python3 scripts/build-sector-targets.py
"""
import hashlib
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl is required: pip install openpyxl')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'public/data/eu-policy-targets-master.xlsx')
OUT = os.path.join(ROOT, 'src/data/sector-targets.generated.ts')

# (identifier fragment, target #) -> (kind, start year, end year, note)
TEMPORAL = {
    ('2023/1804', 1): ('from', 2024, None, None),
    ('2023/1804', 2): ('periodic', 2024, None, 'end of each year'),
    ('2023/1804', 3): ('periodic', 2024, None, 'end of each year'),
    ('2023/1804', 6): ('by', None, 2027, None),
    ('2023/1804', 8): ('by', None, 2027, None),
    ('2023/1804', 9): ('by', None, 2030, None),
    ('2023/1804', 10): ('by', None, 2035, None),
    ('2023/1804', 14): ('by', None, 2027, None),
    ('2023/1804', 15): ('by', None, 2030, None),
    ('2023/1804', 16): ('by', None, 2030, None),
    ('2023/1804', 17): ('by', None, 2027, None),
    ('2023/1804', 18): ('by', None, 2030, None),
    ('2023/1804', 20): ('by', None, 2030, None),
    ('2023/1804', 21): ('by', None, 2030, None),
    ('2023/1804', 22): ('by', None, 2030, None),
    ('2023/1804', 23): ('by', None, 2030, None),
    ('2023/1804', 25): ('by', None, 2029, None),
    ('2023/1804', 26): ('by', None, 2029, None),
    ('2023/1804', 27): ('by', None, 2029, None),
    ('2023/1804', 29): ('by', None, 2029, None),
    ('2023/1804', 31): ('stepped', 2024, 2029, None),
    ('2023/1804', 33): ('by', None, 2029, None),
    ('2023/1804', 34): ('from', 2030, None, None),
    ('2023/1542', 5): ('by', None, 2030, None),
    ('2023/1542', 6): ('by', None, 2030, None),
    ('2021/2115', 13): ('range', 2023, 2027, None),
    ('2021/2115', 14): ('periodic', 2023, None, 'annually'),
    ('2019/631', 1): ('from', 2020, None, None),
    ('2019/631', 2): ('from', 2025, None, None),
    ('2019/631', 3): ('from', 2025, None, None),
    ('2019/631', 4): ('from', 2030, None, None),
    ('2019/631', 5): ('from', 2035, None, None),
    ('2019/631', 6): ('from', 2030, None, None),
    ('2019/631', 7): ('from', 2035, None, None),
    ('2019/631', 8): ('range', 2025, 2029, None),
    ('2021/1060', 2): ('by', None, 2050, None),
    ('COM(2025) 30', 4): ('by', None, 2050, '2040 milestone'),
    ('2024/1252', 2): ('by', None, 2030, None),
    ('2024/1252', 3): ('by', None, 2030, None),
    ('2024/1252', 4): ('by', None, 2030, None),
    ('2024/1252', 5): ('by', None, 2030, None),
    ('2018/842', 1): ('range', 2021, 2030, None),
    ('2018/842', 2): ('by', None, 2030, None),
    ('2022/591', 1): ('by', None, 2050, None),
    ('2022/591', 2): ('by', None, 2030, None),
    ('2022/591', 4): ('by', None, 2030, None),
    ('2022/591', 6): ('by', None, 2030, None),
    ('2022/591', 7): ('by', None, 2030, None),
    ('2023/1791', 1): ('by', None, 2030, None),
    ('2023/1791', 3): ('periodic', 2022, None, 'each year vs 2021'),
    ('2024/1275', 1): ('by', None, 2050, None),
    ('2024/1275', 4): ('stepped', 2030, 2033, None),
    ('2024/1275', 5): ('range', 2026, 2050, None),
    ('2024/1275', 6): ('stepped', 2030, 2050, None),
    ('2024/1275', 12): ('by', None, 2033, None),
    ('COM(2021) 82', 3): ('by', None, 2050, None),
    ('2021/1119', 1): ('by', None, 2050, None),
    ('2021/1119', 3): ('by', None, 2050, None),
    ('2021/1119', 5): ('by', None, 2030, None),
    ('2021/1119', 6): ('by', None, 2030, None),
    ('2021/1119', 7): ('by', None, 2040, None),
    ('2003/87', 3): ('stepped', 2024, 2026, 'ongoing from 2026'),
    ('2003/87', 5): ('from', 2021, None, None),
    ('COM(2019) 640', 5): ('by', None, 2050, None),
    ('COM(2019) 640', 7): ('range', 2021, 2027, None),
    ('2024/573', 11): ('from', 2025, None, None),
    ('2024/573', 27): ('range', 2025, 2028, None),
    ('2024/573', 30): ('from', 2036, None, None),
    ('COM(2021) 550', 2): ('by', None, 2030, None),
    ('COM(2021) 550', 4): ('by', None, 2030, None),
    ('COM(2021) 550', 5): ('by', None, 2030, None),
    ('COM(2021) 550', 6): ('by', None, 2030, None),
    ('COM(2021) 550', 8): ('by', None, 2030, None),
    ('COM(2021) 550', 9): ('by', None, 2050, None),
    ('COM(2021) 550', 10): ('range', 2021, 2027, None),
    ('2023/1805', 1): ('by', None, 2050, None),
    ('2023/1805', 4): ('stepped', 2025, 2050, None),
    ('2018/1999', 7): ('range', 2021, 2030, None),
    ('2018/1999', 8): ('range', 2021, 2030, None),
    ('2018/1999', 9): ('by', None, 2030, None),
    ('2018/1999', 10): ('by', None, 2030, None),
    ('2018/841', 2): ('range', 2026, 2030, None),
    ('2018/841', 4): ('by', None, 2030, None),
    ('2018/841', 5): ('by', None, 2030, None),
    ('2018/841', 10): ('range', 2026, 2030, None),
    ('COM(2024) 91', 1): ('by', None, 2030, None),
    ('2024/1991', 1): ('stepped', 2030, 2050, None),
    ('2024/1991', 3): ('by', None, 2030, None),
    ('2024/1991', 4): ('by', None, 2030, None),
    ('2024/1991', 5): ('stepped', 2030, 2050, None),
    ('2024/1991', 7): ('by', None, 2050, None),
    ('2024/1991', 9): ('by', None, 2030, None),
    ('2024/1991', 10): ('stepped', 2030, 2040, None),
    ('2024/1991', 11): ('by', None, 2050, None),
    ('2024/1991', 12): ('by', None, 2030, None),
    ('2024/1991', 13): ('periodic', 2031, None, 'every 6 years'),
    ('2024/1991', 14): ('by', None, 2030, None),
    ('2024/1991', 16): ('stepped', 2030, 2050, None),
    ('2024/1991', 18): ('periodic', 2030, None, 'every 6 years'),
    ('2024/1991', 19): ('periodic', 2030, None, 'every 6 years'),
    ('2024/1991', 20): ('periodic', 2030, None, 'every 6 years'),
    ('2024/1991', 21): ('by', None, 2030, None),
    ('2024/1735', 2): ('by', None, 2030, None),
    ('2024/1735', 3): ('by', None, 2040, None),
    ('2024/1735', 5): ('by', None, 2030, None),
    ('2024/1735', 6): ('by', None, 2030, None),
    ('2025/40', 7): ('by', None, 2030, None),
    ('2025/40', 8): ('from', 2030, None, None),
    ('2025/40', 9): ('from', 2040, None, None),
    ('2025/40', 12): ('from', 2030, None, None),
    ('2025/40', 13): ('from', 2040, None, None),
    ('2025/40', 14): ('from', 2030, None, None),
    ('2025/40', 15): ('from', 2040, None, None),
    ('2025/40', 16): ('from', 2030, None, None),
    ('2025/40', 19): ('stepped', 2030, 2040, None),
    ('2025/40', 20): ('by', None, 2029, None),
    ('2025/40', 22): ('by', None, 2030, None),
    ('2025/40', 24): ('by', None, 2030, None),
    ('2023/2405', 5): ('from', 2030, None, None),
    ('2023/2405', 6): ('range', 2030, 2031, None),
    ('2023/2405', 7): ('range', 2032, 2034, None),
    ('2023/2405', 8): ('from', 2035, None, None),
    ('2023/2405', 9): ('from', 2040, None, None),
    ('2023/2405', 10): ('from', 2045, None, None),
    ('2023/2405', 11): ('from', 2050, None, None),
    ('2018/2001', 1): ('by', None, 2030, None),
    ('2018/2001', 2): ('by', None, 2030, None),
    ('2018/2001', 3): ('by', None, 2030, None),
    ('2018/2001', 4): ('by', None, 2030, None),
    ('2018/2001', 7): ('range', 2021, 2030, None),
    ('2018/2001', 8): ('stepped', 2030, 2035, None),
    ('2018/2001', 9): ('range', 2021, 2030, None),
    ('2018/2001', 10): ('range', 2021, 2030, None),
    ('2018/2001', 12): ('by', None, 2030, None),
    ('2018/2001', 13): ('stepped', 2025, 2030, None),
    ('2018/2001', 14): ('from', 2030, None, None),
    ('2018/2001', 17): ('range', 2023, 2030, None),
    ('2018/2001', 20): ('from', 2021, None, None),
    ('2018/2001', 21): ('from', 2023, None, None),
    ('COM(2020) 662', 2): ('by', None, 2030, None),
    ('COM(2020) 662', 3): ('range', 2030, 2050, None),
    ('COM(2020) 662', 4): ('by', None, 2030, None),
    ('2019/904', 1): ('by', None, 2026, None),
    ('2019/904', 4): ('from', 2025, None, None),
    ('2019/904', 5): ('from', 2030, None, None),
    ('2019/904', 8): ('by', None, 2029, None),
    ('2022/869', 1): ('by', None, 2050, None),
    ('2024/1679', 22): ('by', None, 2030, None),
    ('2008/98', 2): ('by', None, 2030, None),
    ('2008/98', 6): ('by', None, 2030, None),
    ('2008/98', 7): ('by', None, 2035, None),
    ('COM(2025) 280', 3): ('by', None, 2030, None),
    ('COM(2025) 280', 4): ('by', None, 2030, None),
    ('COM(2025) 280', 7): ('by', None, 2030, None),
    ('COM(2025) 280', 8): ('by', None, 2030, None),
    ('COM(2025) 280', 9): ('by', None, 2030, None),
    ('COM(2025) 280', 10): ('by', None, 2027, None),
    ('COM(2025) 280', 12): ('range', 2026, 2050, None),
    ('COM(2025) 280', 13): ('by', None, 2030, None),
    ('COM(2025) 280', 14): ('by', None, 2027, None),
    ('COM(2025) 280', 19): ('by', None, 2030, None),
    ('COM(2025) 280', 20): ('by', None, 2033, None),
    ('COM(2021) 400', 2): ('by', None, 2030, None),
    ('COM(2021) 400', 3): ('by', None, 2030, None),
    ('COM(2021) 400', 4): ('by', None, 2030, None),
    ('COM(2021) 400', 5): ('by', None, 2030, None),
    ('COM(2021) 400', 6): ('by', None, 2030, None),
    ('COM(2021) 400', 7): ('by', None, 2030, None),
    ('COM(2021) 400', 8): ('by', None, 2030, None),
    ('COM(2021) 400', 9): ('by', None, 2030, None),
}

# identifier fragment -> compact display name
POLICY_SHORT = {
    '2023/1804': 'AFIR (alternative fuels infrastructure)',
    '2023/1542': 'Batteries Regulation',
    '2021/2115': 'CAP Strategic Plans Regulation',
    '2019/631': 'CO2 standards for cars and vans',
    '2021/1060': 'Common Provisions Regulation / JTF',
    'COM(2025) 30': 'Competitiveness Compass',
    '2024/1252': 'Critical Raw Materials Act',
    '2018/842': 'Effort Sharing Regulation',
    '2022/591': '8th Environment Action Programme',
    '2023/1791': 'Energy Efficiency Directive',
    '2024/1275': 'Energy Performance of Buildings Directive',
    'COM(2021) 82': 'EU Adaptation Strategy',
    '2021/1119': 'European Climate Law',
    '2003/87': 'EU ETS Directive',
    'COM(2019) 640': 'European Green Deal',
    '2024/573': 'F-gas Regulation',
    'COM(2021) 550': 'Fit for 55 Communication',
    '2023/1805': 'FuelEU Maritime',
    '2018/1999': 'Governance Regulation',
    '2018/841': 'LULUCF Regulation',
    'COM(2024) 91': 'Climate Risks Communication',
    '2024/1991': 'Nature Restoration Regulation',
    '2024/1735': 'Net-Zero Industry Act',
    '2025/40': 'Packaging Regulation (PPWR)',
    '2023/2405': 'ReFuelEU Aviation',
    '2018/2001': 'Renewable Energy Directive (RED III)',
    'COM(2020) 662': 'Renovation Wave',
    '2019/904': 'Single-Use Plastics Directive',
    '2022/869': 'TEN-E Regulation',
    '2024/1679': 'TEN-T Regulation',
    '2008/98': 'Waste Framework Directive',
    'COM(2025) 280': 'Water Resilience Strategy',
    'COM(2021) 400': 'Zero Pollution Action Plan',
}
FRAGS = sorted(POLICY_SHORT, key=len, reverse=True)

SECTOR_COLS = [
    ('energy', 22), ('buildings', 23), ('agrifood', 24), ('transport', 25),
    ('industry', 26), ('ecosystems', 27), ('water', 28), ('health', 29),
]
CLIMATE = {
    'Mitigation': 'mitigation', 'Adaptation': 'adaptation',
    'Mitigation + Adaptation': 'both', 'Not climate-specific': 'none',
}


def frag_of(name):
    for f in FRAGS:
        if f in name:
            return f
    raise KeyError('no policy fragment matches: ' + name)


def fallback_temporal(tl):
    years = [int(y) for y in re.findall(r'20\d\d', tl or '')]
    if not years:
        return ('by', None, None, None)
    if len(years) == 1:
        return ('from' if re.match(r'(?i)\s*from', tl or '') else 'by',
                years[0] if re.match(r'(?i)\s*from', tl or '') else None,
                None if re.match(r'(?i)\s*from', tl or '') else years[0], None)
    return ('range', min(years), max(years), None)


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb['Clean targets']
rows = []
for r in range(2, ws.max_row + 1):
    cell = lambda c: ws.cell(row=r, column=c).value
    name = cell(1)
    if not name:
        continue
    frag = frag_of(str(name))
    num = int(cell(5))
    key = (frag, num)
    kind, start, end, tnote = TEMPORAL.get(key) or fallback_temporal(str(cell(12) or ''))
    note = (cell(33) or '').strip() if cell(33) else ''
    indicators = [s.strip() for s in str(cell(14) or '').split(';') if s.strip() and s.strip().lower() != 'none']
    rows.append({
        'id': hashlib.sha1(f'{name}#{num}'.encode()).hexdigest()[:12],
        'policy': str(name).strip(),
        'policyShort': POLICY_SHORT[frag],
        'docType': str(cell(2) or '').strip(),
        'area': str(cell(4) or '').strip(),
        'num': num,
        'order': int(cell(6)) if cell(6) else 1,
        'text': str(cell(7) or '').strip(),
        'short': str(cell(32) or '').strip(),
        'provision': str(cell(8) or '').strip(),
        'label': str(cell(9) or '').strip().lower(),
        'obligation': str(cell(10) or '').strip().lower(),
        'targetType': str(cell(11) or '').strip().lower(),
        'timeline': str(cell(12) or '').strip(),
        'indicators': indicators,
        'climate': CLIMATE.get(str(cell(15) or '').strip(), 'none'),
        'source': str(cell(16) or '').strip(),
        'sectors': [k for k, c in SECTOR_COLS if str(cell(c) or '').strip() == 'Yes'],
        'climateArgument': str(cell(30) or '').strip(),
        'revisionNote': note,
        'crossRef': 'cross-reference' in note.lower(),
        'temporal': {'kind': kind, 'start': start, 'end': end, 'note': tnote},
    })

if not all(t['short'] for t in rows):
    sys.exit('rows missing a short version — fix the Masterfile first')

header = (
    '// GENERATED by scripts/build-sector-targets.py — do not edit by hand.\n'
    '// Source: public/data/eu-policy-targets-master.xlsx, tab "Clean targets"\n'
    f'// ({len(rows)} rows). Regenerate after any Masterfile revision.\n'
    'import type { SectorTarget } from \'./sector-targets\';\n\n'
    'export const RAW_SECTOR_TARGETS: SectorTarget[] = '
)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(header + json.dumps(rows, ensure_ascii=False, indent=1) + ';\n')
print(f'wrote {OUT} with {len(rows)} rows')
