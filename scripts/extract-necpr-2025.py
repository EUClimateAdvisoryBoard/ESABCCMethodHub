#!/usr/bin/env python3
"""Extract tidy datasets from the 2025 NECPR reporting workbooks (EEA Reportnet 3).

Reads the three published annex workbooks and writes flat CSVs plus a
data-quality report to public/data/necpr-2025/.

    python3 scripts/extract-necpr-2025.py \
        --ghg  NECPR_Annex_I_GHG_Progress_2025.xlsx \
        --res  NECPR_Annex_II_RES_2025.xlsx \
        --ee   NECPR_Annex_IV_EE_Progress_2025.xlsx

Outputs
    annex1-table1-targets.csv     national 2030/2040/2050 targets, projections, gap
    annex1-table1-neutrality.csv  climate-neutrality year + role of removals
    annex1-table3-lulucf.csv      LULUCF history, projections, NECP commitment
    annex1-table4-other.csv       other national GHG targets (one row per target-year)
    annex2-table7-res.csv         other RES objectives
    annex4-table6-ee.csv          other EE objectives (free text only)
    quality-flags.csv             machine-detected data-quality flags

The workbooks are not redistributed in this repository; pass the paths of the
files downloaded from the EEA.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from collections import defaultdict

import openpyxl

OUT_DIR = os.path.join("public", "data", "necpr-2025")

# Scope labels used in Annex I Table 1, shortened for the tidy output.
SCOPES = {
    "Total GHG emissions excluding LULUCF, excluding international aviation": "excl_lulucf",
    "Total GHG emissions including LULUCF, excluding international aviation": "incl_lulucf",
    "Total GHG including LULUCF, including international aviation": "incl_lulucf_aviation",
}

NOTATION_KEYS = {"NA", "NAV", "NAv", "N/A", "NO", "ΝΑ", "brak danych", "Nie dotyczy", ""}


def sheet_rows(wb, name):
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    header = [c for c in next(it)]
    out = []
    for row in it:
        if not any(c is not None for c in row):
            continue
        out.append({h: v for h, v in zip(header, row) if h is not None})
    return out


def num(value):
    """Numeric value, or None for notation keys and free text."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in NOTATION_KEYS:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def write_csv(name, fieldnames, rows):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"  {path}  ({len(rows)} rows)")
    return path


# --------------------------------------------------------------------------
# Annex I — GHG
# --------------------------------------------------------------------------

def annex1(path, flags):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ---- Table 1: targets, projections, gap ------------------------------
    element = {
        "National GHG target (for 2030 and beyond)": "target",
        "Article 18 WEM scenario": "wem",
        "Article 18 WAM scenario": "wam",
        "Historic emissions": "historic",
    }
    cube = defaultdict(dict)
    neutrality = {}
    removals = defaultdict(dict)

    for row in sheet_rows(wb, "Table_1"):
        cc, el = row["countryCode"], row["Reporting_element"]
        if el == "Climate neutrality":
            neutrality[cc] = row["Target_year_for_climate_neutrality"]
            continue
        if el == "Role of removals":
            # The template reuses the neutrality-year column for the NECP value.
            removals[cc][str(row["year"])] = (
                num(row["value"]),
                num(row["Target_year_for_climate_neutrality"]),
            )
            continue
        key = element.get(el)
        if key is None:
            if "Difference between WEM" in str(el):
                key = "gap_wem_rel"
            elif "Difference between WAM" in str(el):
                key = "gap_wam_rel"
            else:
                continue
        scope = SCOPES.get(row["Scope"])
        if scope is None:
            continue
        cube[(cc, scope, str(row["year"]))][key] = num(row["value"])

    t1 = []
    for (cc, scope, year), vals in sorted(cube.items()):
        target, wem, wam = vals.get("target"), vals.get("wem"), vals.get("wam")
        t1.append(
            {
                "country": cc,
                "scope": scope,
                "year": year,
                "unit": "kt CO2e",
                "target": target,
                "historic": vals.get("historic"),
                "wem": wem,
                "wam": wam,
                "gap_wem_kt": None if (target is None or wem is None) else wem - target,
                "gap_wam_kt": None if (target is None or wam is None) else wam - target,
                "gap_wem_rel_reported": vals.get("gap_wem_rel"),
                "gap_wam_rel_reported": vals.get("gap_wam_rel"),
            }
        )
        # A target more than 3x the country's own WEM projection is not a target.
        if target and wem and target > 3 * wem:
            flags.append(
                {
                    "dataset": "Annex I Table 1",
                    "country": cc,
                    "item": f"{scope} {year} national target",
                    "flag": "implausible_magnitude",
                    "detail": f"target {target:,.0f} kt vs own WEM projection {wem:,.0f} kt "
                    f"(ratio {target / wem:.0f}x) — order-of-magnitude entry error",
                }
            )

    write_csv(
        "annex1-table1-targets.csv",
        [
            "country", "scope", "year", "unit", "target", "historic", "wem", "wam",
            "gap_wem_kt", "gap_wam_kt", "gap_wem_rel_reported", "gap_wam_rel_reported",
        ],
        t1,
    )

    neut = []
    for cc in sorted(set(neutrality) | set(removals)):
        year = neutrality.get(cc)
        row = {
            "country": cc,
            "climate_neutrality_year": None if str(year) in NOTATION_KEYS else year,
            "climate_neutrality_reported": str(year) not in NOTATION_KEYS,
        }
        for y in ("2030", "2040", "2050"):
            value, necp = removals.get(cc, {}).get(y, (None, None))
            row[f"removals_{y}_kt"] = value
            row[f"removals_{y}_necp_kt"] = necp
        neut.append(row)
        if str(year) in NOTATION_KEYS:
            flags.append(
                {
                    "dataset": "Annex I Table 1",
                    "country": cc,
                    "item": "climate neutrality year",
                    "flag": "not_reported",
                    "detail": "no target year given for climate neutrality",
                }
            )

    write_csv(
        "annex1-table1-neutrality.csv",
        ["country", "climate_neutrality_year", "climate_neutrality_reported"]
        + [f"removals_{y}_{s}" for y in ("2030", "2040", "2050") for s in ("kt", "necp_kt")],
        neut,
    )

    # ---- Table 3: LULUCF --------------------------------------------------
    t3_element = {
        "Land Use, Land-Use Change and Forestry": "historic",
        "Land Use, Land-Use Change and Forestry in the WEM scenario": "wem",
        "Land Use, Land-Use Change and Forestry in the WAM scenario": "wam",
        "LULUCF commitment stated in current NECP": "necp",
    }
    lulucf = defaultdict(dict)
    units = defaultdict(set)
    for row in sheet_rows(wb, "Table_3"):
        key = t3_element.get(row["Reporting_element"])
        if key is None:
            continue
        lulucf[(row["countryCode"], str(row["year"]))][key] = num(row["value"])
        if row["unit"]:
            units[row["countryCode"]].add(str(row["unit"]).strip())

    t3 = []
    for (cc, year), vals in sorted(lulucf.items()):
        if year == "None":
            continue
        t3.append(
            {
                "country": cc,
                "year": year,
                "unit": "kt CO2e",
                "historic": vals.get("historic"),
                "wem": vals.get("wem"),
                "wam": vals.get("wam"),
                "necp_commitment": vals.get("necp"),
            }
        )

    for cc in sorted({k[0] for k in lulucf}):
        vals = lulucf.get((cc, "2030"), {})
        necp, wem, wam = vals.get("necp"), vals.get("wem"), vals.get("wam")
        for label, proj in (("WAM", wam), ("WEM", wem)):
            if necp is not None and proj is not None and abs(necp - proj) < max(1.0, abs(proj) * 0.005):
                flags.append(
                    {
                        "dataset": "Annex I Table 3",
                        "country": cc,
                        "item": "2030 LULUCF NECP commitment",
                        "flag": "commitment_equals_projection",
                        "detail": f"stated commitment {necp:,.1f} kt is the country's own "
                        f"{label} projection, not an independent commitment",
                    }
                )
                break
        # A commitment three orders of magnitude off the projection is a unit error.
        if necp is not None and wem is not None and necp != 0 and abs(wem) > 100:
            ratio = abs(wem) / max(abs(necp), 1e-9)
            if ratio > 100:
                flags.append(
                    {
                        "dataset": "Annex I Table 3",
                        "country": cc,
                        "item": "2030 LULUCF NECP commitment",
                        "flag": "unit_mismatch",
                        "detail": f"commitment {necp:g} vs projection {wem:,.0f} kt — "
                        f"commitment appears to be in Mt while the column is kt",
                    }
                )

    write_csv(
        "annex1-table3-lulucf.csv",
        ["country", "year", "unit", "historic", "wem", "wam", "necp_commitment"],
        t3,
    )

    for cc, seen in sorted(units.items()):
        others = {u for u in seen if u.replace("\xa0", "").strip() not in ("kt CO2e", "ktCO2e")}
        if others:
            flags.append(
                {
                    "dataset": "Annex I Table 3",
                    "country": cc,
                    "item": "unit field",
                    "flag": "nonstandard_unit_string",
                    "detail": "reported unit(s): " + ", ".join(sorted(others)),
                }
            )

    # ---- Table 4: other national GHG targets ------------------------------
    blocks, current = [], None
    for row in sheet_rows(wb, "Table_4"):
        el = row["Reporting_element"]
        if el == "Target/ objective":
            key = (
                row["countryCode"],
                row["Name_of_national_target_or_objective"],
                row["Sectors_addressed"],
                row["unit"],
            )
            if current is None or current["key"] != key:
                current = {
                    "key": key,
                    "description": row["Description"],
                    "gwp": row["GWP_used"],
                    "target": {},
                    "current": {},
                    "wem": {},
                    "wam": {},
                }
                blocks.append(current)
            current["target"][str(row["year"])] = row["value"]
        elif current is not None:
            series = {
                "Current progress": "current",
                "Projected progress under WEM scenario": "wem",
                "Projected progress under WAM scenario": "wam",
            }.get(el)
            if series:
                current[series][str(row["year"])] = row["value"]

    t4 = []
    for block in blocks:
        cc, name, sector, unit = block["key"]
        numeric_target_years = [y for y, v in block["target"].items() if num(v) is not None]
        for year in ("2022", "2023", "2030", "2035", "2040"):
            t4.append(
                {
                    "country": cc,
                    "target_name": name,
                    "sector": sector,
                    "unit": unit,
                    "gwp": block["gwp"],
                    "year": year,
                    "target": block["target"].get(year),
                    "current": block["current"].get(year),
                    "wem": block["wem"].get(year),
                    "wam": block["wam"].get(year),
                    "target_numeric": num(block["target"].get(year)),
                    "has_any_numeric_target": bool(numeric_target_years),
                    "description": (block["description"] or "").strip(),
                }
            )
        if not numeric_target_years:
            flags.append(
                {
                    "dataset": "Annex I Table 4",
                    "country": cc,
                    "item": str(name)[:80],
                    "flag": "no_numeric_target",
                    "detail": "objective reported with no quantified target in any year",
                }
            )
        if str(unit).strip().lower() in ("brak danych", "nie dotyczy", "na", "none"):
            flags.append(
                {
                    "dataset": "Annex I Table 4",
                    "country": cc,
                    "item": str(name)[:80],
                    "flag": "missing_unit",
                    "detail": f"unit field reads '{unit}'",
                }
            )
        # Percentage targets stated as fractions alongside absolute progress series.
        if str(unit).strip() == "%":
            progress = [num(v) for v in block["current"].values() if num(v) is not None]
            targets = [num(v) for v in block["target"].values() if num(v) is not None]
            if targets and progress and max(abs(t) for t in targets) <= 1 < max(abs(p) for p in progress):
                flags.append(
                    {
                        "dataset": "Annex I Table 4",
                        "country": cc,
                        "item": str(name)[:80],
                        "flag": "fraction_vs_percent",
                        "detail": "target given as a fraction (e.g. -0.7) while the progress "
                        "series is in percent or absolute units",
                    }
                )

    write_csv(
        "annex1-table4-other.csv",
        [
            "country", "target_name", "sector", "unit", "gwp", "year", "target",
            "current", "wem", "wam", "target_numeric", "has_any_numeric_target",
            "description",
        ],
        t4,
    )

    # ---- Table 2 unit check ----------------------------------------------
    aea, esr = {}, {}
    for row in sheet_rows(wb, "Table_2"):
        if row["Reporting_element"] == "Annual emission allocation":
            aea[(row["countryCode"], str(row["year"]))] = (num(row["value"]), row["unit"])
        elif row["Reporting_element"] == "Total Effort Sharing emissions - WEM":
            esr[(row["countryCode"], str(row["year"]))] = num(row["value"])
    for key, (value, unit) in sorted(aea.items()):
        other = esr.get(key)
        if value and other and other / value > 100:
            flags.append(
                {
                    "dataset": "Annex I Table 2",
                    "country": key[0],
                    "item": f"{key[1]} annual emission allocation",
                    "flag": "unit_mismatch",
                    "detail": f"AEA {value:,.1f} labelled '{unit}' but ESR emissions for the "
                    f"same year are {other:,.0f} kt — the AEA series is in Mt, and the "
                    f"pre-calculated difference column subtracts Mt from kt",
                }
            )
    wb.close()


# --------------------------------------------------------------------------
# Annex II — RES
# --------------------------------------------------------------------------

STANDARD_RES_OBJECTIVES = [
    "Renewable energy use in district heating",
    "Renewable energy use in buildings",
    "Renewable energy produced by cities",
    "Renewable energy communities",
    "Renewables self-consumers",
    "Energy recovered from the sludge acquired through the treatment of wastewater",
]


def annex2(path, flags):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = sheet_rows(wb, "Table_7")
    out = []
    units_by_objective = defaultdict(set)
    for row in rows:
        objective = str(row["Trajectory_or_objective"])
        standardised = objective in STANDARD_RES_OBJECTIVES
        target_num = num(row["Target"])
        value_num = num(row["value"])
        out.append(
            {
                "country": row["countryCode"],
                "objective": objective,
                "objective_standardised": standardised,
                "target_text": row["Target"],
                "target_numeric": target_num,
                "target_year": row["Target_year"],
                "indicator": row["Name_of_indicator_to_monitor_progress"],
                "unit": row["Unit"],
                "year": row["year"],
                "value": value_num,
                "comparable": target_num is not None and value_num is not None,
                "description": (row["Description"] or "").strip() if row["Description"] else "",
            }
        )
        if standardised and value_num is not None and row["Unit"]:
            units_by_objective[objective].add(str(row["Unit"]).strip())
        if target_num is None and str(row["Target"]).strip() not in NOTATION_KEYS and row["Target"]:
            flags.append(
                {
                    "dataset": "Annex II Table 7",
                    "country": row["countryCode"],
                    "item": objective[:80],
                    "flag": "target_as_free_text",
                    "detail": f"target field is prose, not a number: "
                    f"\"{str(row['Target'])[:120]}\"",
                }
            )

    for objective, units in sorted(units_by_objective.items()):
        if len(units) > 1:
            flags.append(
                {
                    "dataset": "Annex II Table 7",
                    "country": "(all)",
                    "item": objective[:80],
                    "flag": "incomparable_units",
                    "detail": f"{len(units)} different units used for one standardised "
                    f"objective: " + ", ".join(sorted(units)),
                }
            )

    write_csv(
        "annex2-table7-res.csv",
        [
            "country", "objective", "objective_standardised", "target_text",
            "target_numeric", "target_year", "indicator", "unit", "year", "value",
            "comparable", "description",
        ],
        out,
    )
    wb.close()


# --------------------------------------------------------------------------
# Annex IV — EE
# --------------------------------------------------------------------------

NUMBER_IN_TEXT = re.compile(r"-?\d[\d\s.,]*\s*(?:%|TWh|GWh|PJ|ktoe|Mtoe|Mtep|MW|GW|kt|Mt|toe)\b", re.I)


def annex4(path, flags):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = sheet_rows(wb, "Table_6")
    out = []
    for row in rows:
        name = row["Name_of_national_target_or_objective"]
        empty = name is None or str(name).strip() in NOTATION_KEYS
        blob = " ".join(
            str(row.get(c) or "")
            for c in ("Description", "Progress_towards_target_or_objective", "Expected_impacts_of_the_set_objective")
        )
        quantities = NUMBER_IN_TEXT.findall(blob)
        out.append(
            {
                "country": row["countryCode"],
                "target_name": None if empty else str(name).strip(),
                "reported": not empty,
                "description": (row["Description"] or "").strip() if row["Description"] else "",
                "progress": (row["Progress_towards_target_or_objective"] or "").strip()
                if row["Progress_towards_target_or_objective"] else "",
                "expected_impact": (row["Expected_impacts_of_the_set_objective"] or "").strip()
                if row["Expected_impacts_of_the_set_objective"] else "",
                "quantities_in_text": len(quantities),
            }
        )
        if empty:
            flags.append(
                {
                    "dataset": "Annex IV Table 6",
                    "country": row["countryCode"],
                    "item": "(blank row)",
                    "flag": "empty_row",
                    "detail": "row submitted with NA in every field",
                }
            )
        elif not quantities:
            flags.append(
                {
                    "dataset": "Annex IV Table 6",
                    "country": row["countryCode"],
                    "item": str(name)[:80],
                    "flag": "no_quantity_anywhere",
                    "detail": "no numeric value with a unit appears in any free-text field",
                }
            )

    flags.append(
        {
            "dataset": "Annex IV Table 6",
            "country": "(all)",
            "item": "table structure",
            "flag": "no_value_columns",
            "detail": "the table has no unit, year or value column — every objective, its "
            "target and its progress are free text, so nothing is machine-comparable",
        }
    )

    write_csv(
        "annex4-table6-ee.csv",
        [
            "country", "target_name", "reported", "description", "progress",
            "expected_impact", "quantities_in_text",
        ],
        out,
    )
    wb.close()


def main():
    global OUT_DIR
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--ghg", required=True, help="NECPR_Annex_I_GHG_Progress_2025.xlsx")
    parser.add_argument("--res", required=True, help="NECPR_Annex_II_RES_2025.xlsx")
    parser.add_argument("--ee", required=True, help="NECPR_Annex_IV_EE_Progress_2025.xlsx")
    parser.add_argument("--out", default=OUT_DIR)
    args = parser.parse_args()

    OUT_DIR = args.out
    os.makedirs(OUT_DIR, exist_ok=True)

    flags = []
    print("Annex I (GHG):")
    annex1(args.ghg, flags)
    print("Annex II (RES):")
    annex2(args.res, flags)
    print("Annex IV (EE):")
    annex4(args.ee, flags)

    print("Data-quality flags:")
    write_csv("quality-flags.csv", ["dataset", "country", "item", "flag", "detail"], flags)


if __name__ == "__main__":
    main()
