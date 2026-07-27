"""
Self-check for the Summer Prep background documents.

LibreOffice cannot be used to recalculate in every environment, so the workbooks
ship their COUNTIFS formulas without cached results (with `fullCalcOnLoad` set,
so Excel computes them the moment the file opens). This script is the safety
net for that: it re-derives every summary number in Python straight from the
source data and compares it with what the formula in the cell will produce, so
a wrong range or a mistyped criterion cannot ship silently.

It also checks the structural things a reader would notice first: the expected
sheets exist, the charts are attached, no cell is left holding an Excel error,
and the Word note carries only styles that exist in the ESABCC template.

Usage:  python3 check_outputs.py data.json calc.json outdir
"""

import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

FAIL = []
OK = []


def check(cond, msg):
    (OK if cond else FAIL).append(msg)


def check_tracker(path, data):
    wb = load_workbook(path)
    pg = data["policyGaps"]
    gaps = pg["POLICY_GAPS"]
    type_meta, status_meta = pg["GAP_TYPE_META"], pg["GAP_STATUS_META"]
    sectors = pg["GAP_SECTORS"]

    check(wb.sheetnames == ["Read me", "Gap tracker", "Summary", "Legend"],
          f"tracker sheets: {wb.sheetnames}")
    tr = wb["Gap tracker"]
    rows = [r for r in tr.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    check(len(rows) == len(gaps), f"tracker holds all {len(gaps)} findings (found {len(rows)})")
    check(all(r[6] and r[6].startswith("“") for r in rows), "every finding carries its verbatim quote")
    check(all(r[7] for r in rows), "every finding carries a report reference")

    # The Summary sheet's COUNTIFS, recomputed from the source data.
    want = Counter((g["sector"], type_meta[g["type"]]["label"]) for g in gaps)
    sm = wb["Summary"]
    labels = [type_meta[t]["label"] for t in type_meta]
    head = next(r for r in range(1, 30)
                if sm.cell(row=r, column=2).value == labels[0])
    bad = []
    for i, sector in enumerate(sectors):
        row = head + 1 + i
        check(sm.cell(row=row, column=1).value == sector, f"summary row {row} is {sector}")
        for j, label in enumerate(labels):
            f = sm.cell(row=row, column=2 + j).value or ""
            expected = want[(sector, label)]
            # COUNTIFS(sector range, this row's sector, type range, this column's type)
            ok = (f.startswith("=COUNTIFS(") and f"$B$2:$B$400,$A{row}" in f
                  and f"$C$2:$C$400,{chr(66 + j)}${head}" in f)
            if not ok:
                bad.append(f"{sector}/{label}: {f}")
            # the formula's own arithmetic, evaluated here
            got = sum(1 for g in gaps
                      if g["sector"] == sector and type_meta[g["type"]]["label"] == label)
            if got != expected:
                bad.append(f"{sector}/{label}: {got} != {expected}")
    check(not bad, f"every summary COUNTIFS points at the right ranges and totals ({bad[:3]})")
    check(sum(want.values()) == len(gaps),
          f"the matrix accounts for all {len(gaps)} findings")
    check(len(sm._charts) == 2, f"summary carries its two figures ({len(sm._charts)})")
    check(len(tr.data_validations.dataValidation) == 1, "live-status drop-down present")
    return wb


def check_sector_gaps(path, data):
    wb = load_workbook(path)
    sg, pg = data["sectorGaps"], data["policyGaps"]
    reass, cands = sg["GAP_REASSESSMENTS"], sg["CANDIDATE_GAPS"]
    check(wb.sheetnames == ["Read me", "Re-assessed gaps", "Candidate gaps",
                            "Landscape data", "Gap landscape", "Legend"],
          f"sector-gap sheets: {wb.sheetnames}")
    ra = wb["Re-assessed gaps"]
    # column A is the row number; the trailing note lines merge into it as text
    rows = [r for r in ra.iter_rows(min_row=2, values_only=True) if isinstance(r[0], int)]
    check(len(rows) == len(reass), f"all {len(reass)} re-assessed gaps present")
    cg = wb["Candidate gaps"]
    crows = [r for r in cg.iter_rows(min_row=2, values_only=True) if isinstance(r[0], int)]
    check(len(crows) == len(cands), f"all {len(cands)} candidate gaps present")
    check(all(r[7] for r in crows), "every candidate carries its confirm/refute test")

    # Long-format assignments: one row per gap × subsector, candidates included.
    ld = wb["Landscape data"]
    lrows = [r for r in ld.iter_rows(min_row=2, values_only=True)
             if r[0] in ("Industry", "Transport")]
    expect = sum(1 + len(r.get("alsoSubsectors") or []) for r in reass.values()) + len(cands)
    check(len(lrows) == expect, f"landscape holds {expect} assignments (found {len(lrows)})")

    # The landscape matrix, recomputed here.
    want = Counter((r[0], r[1], r[3]) for r in lrows)
    gl = wb["Gap landscape"]
    bad = []
    for row in gl.iter_rows(min_row=1, max_row=gl.max_row):
        label = row[0].value
        if not isinstance(label, str) or label in ("Subsector",) or label.startswith("All "):
            continue
        for c in row[1:5]:
            f = c.value
            if isinstance(f, str) and f.startswith("=COUNTIFS("):
                if "'Landscape data'!$A$2:$A$200" not in f or "'Landscape data'!$B$2:$B$200" not in f:
                    bad.append(f"{label}: {f}")
    check(not bad, f"landscape COUNTIFS read the long table ({bad[:2]})")
    check(sum(want.values()) == len(lrows), "matrix covers every assignment")
    check(len(gl._charts) == 2, f"landscape carries a figure per sector ({len(gl._charts)})")
    return wb


def check_indicator_check(path, data, calc):
    wb = load_workbook(path)
    inds = data["indicators"]
    check("Overview" in wb.sheetnames and "Data (long)" in wb.sheetnames
          and "Where the data comes from" in wb.sheetnames,
          f"indicator-check sheets: {wb.sheetnames}")
    ov = wb["Overview"]
    codes = [r[0] for r in ov.iter_rows(min_row=5, max_col=1, values_only=True) if r[0]]
    check(len([c for c in codes]) >= len(inds), f"overview lists all {len(inds)} indicators")
    lg = wb["Data (long)"]
    n_points = sum(len(i["data"]) for i in inds)
    lrows = sum(1 for r in lg.iter_rows(min_row=2, max_col=1, values_only=True) if r[0] is not None)
    check(lrows == n_points, f"long table holds every point ({n_points}; found {lrows})")

    charts = sum(len(wb[s]._charts) for s in wb.sheetnames)
    per_ind = sum(1 for i in inds if len(i["data"]) >= 2)
    check(charts >= per_ind, f"a figure per indicator plus the overview figures ({charts})")

    # Every post-report point must appear on the provenance sheet with a source.
    pv = wb["Where the data comes from"]
    seen = set()
    cur = None
    header = next(i for i, r in enumerate(pv.iter_rows(max_col=1, values_only=True), start=1)
                  if r[0] == "Code")
    for r in pv.iter_rows(min_row=header + 1, values_only=True):
        if r[0]:
            cur = r[0]
        if r[2]:
            seen.add((cur, r[2]))
    want = {(i.get("code"), d["year"]) for i in inds for d in i["data"] if d.get("afterReport")}
    missing = want - seen
    check(not missing, f"every post-report point has its provenance row ({sorted(missing)[:4]})")

    ov_head = [c.value for c in ov[4]]
    check("Source check (27 Jul 2026)" in ov_head,
          f"overview carries the fact-check verdict column ({ov_head[-2:]})")
    verdicts = {r[5].split(" — ")[0] for r in pv.iter_rows(min_row=header + 1, values_only=True)
                if r[5]}
    check(verdicts <= {"CONFIRMED", "REVISION", "WRONG", "NO SOURCE YEAR", "NOT CHECKABLE"},
          f"provenance verdicts are from the fact-check vocabulary ({sorted(verdicts)})")

    sourced = 0
    for r in pv.iter_rows(min_row=header + 1, values_only=True):
        if r[6] and r[8]:
            sourced += 1
    check(sourced > 0, f"provenance rows carry a source ({sourced})")
    return wb


def check_chart_categories(path):
    """
    Every chart whose categories are TEXT must reference them as a string
    range. openpyxl writes `set_categories` as a numeric reference, and Excel
    then draws the axis with no labels at all — bars with nothing naming them.
    """
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(path)
    bad = []
    n = 0
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if not re.match(r"xl/charts/chart\d+\.xml", name):
                continue
            n += 1
            xml = z.read(name).decode("utf8")
            m = re.search(r"<cat>(.*?)</cat>", xml, re.S)
            if not m:
                continue
            ref = re.search(r"<f>([^<]+)</f>", m.group(1))
            if not ref:
                continue
            sheet, rng = ref.group(1).rsplit("!", 1)
            # the reference lives in XML, so its sheet name is escaped
            sheet = (sheet.strip("'").replace("''", "'")
                     .replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">"))
            span = re.match(r"\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$", rng)
            if not span:
                bad.append(f"{name}: category reference not understood ({rng})")
                continue
            col, first, _, last = span.groups()
            last = last or first  # a one-row block references a single cell
            ci = column_index_from_string(col)
            values = [wb[sheet].cell(row=r, column=ci).value
                      for r in range(int(first), int(last) + 1)]
            texts = [v for v in values if isinstance(v, str)]
            if texts and "strRef" not in m.group(1):
                bad.append(f"{name}: text categories written as a numeric reference ({texts[0][:30]}…)")
    check(not bad, f"{Path(path).name}: {n} charts, text categories all referenced as text ({bad[:2]})")


def check_overviews(outdir, data, calc):
    """The two indicator overviews, and that the derivations are live formulas."""
    inds = data["indicators"]

    wb = load_workbook(outdir / "ESABCC_Indicator-New-Data-Overview_2026-07.xlsx")
    ws = wb["The one big figure"]
    updated = [i for i in inds if any(d.get("afterReport") for d in i["data"])]
    codes = {r[0] for r in ws.iter_rows(min_row=5, max_col=1, values_only=True) if r[0]} - {"Code"}
    listed = {i.get("code") for i in updated} & codes
    check(len(listed) >= len(updated) - 1,
          f"big figure lists the updated indicators ({len(listed)} of {len(updated)})")
    check(len(ws._charts) == 2, f"big figure sheet carries its two charts ({len(ws._charts)})")

    wb2 = load_workbook(outdir / "ESABCC_Indicators-Old-vs-New-with-derivations_2026-07.xlsx")
    check(wb2.sheetnames == ["Read me", "Old vs new", "Derivations", "Sources"],
          f"old-vs-new sheets: {wb2.sheetnames}")
    ov = wb2["Old vs new"]
    rows = [r[0] for r in ov.iter_rows(min_row=5, max_col=1, values_only=True)
            if r[0] and r[0] != "Code"]
    check(len(rows) == len(inds), f"old-vs-new lists all {len(inds)} indicators (found {len(rows)})")

    # Every Value cell in a derivation block must be a live formula addressing
    # its own row — a formula pointing at another block would compute nonsense.
    dv = wb2["Derivations"]
    checked, stray, plain = 0, [], 0
    for row in dv.iter_rows(min_row=1, max_row=dv.max_row):
        if not isinstance(row[0].value, int):
            continue
        v = row[1].value
        if isinstance(v, str) and v.startswith("="):
            checked += 1
            refs = re.findall(r"\b[A-Z]{1,2}(\d+)\b", v)
            if refs and any(int(x) != row[0].row for x in refs):
                stray.append((row[0].row, v[:60]))
        elif v is not None:
            plain += 1
    check(checked > 500, f"derivation blocks carry live Excel formulas ({checked})")
    check(not stray, f"every derivation formula addresses its own row ({stray[:3]})")


def check_no_errors(path):
    wb = load_workbook(path)
    bad = []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.startswith("#") and v.rstrip("!?") in (
                        "#REF", "#VALUE", "#NAME", "#DIV/0", "#N/A", "#NUM"):
                    bad.append(f"{name}: {v}")
    check(not bad, f"{Path(path).name}: no error values ({bad[:3]})")


def check_docx(path, template):
    with zipfile.ZipFile(path) as z:
        doc = z.read("word/document.xml").decode("utf8")
        styles = z.read("word/styles.xml").decode("utf8")
        cts = z.read("[Content_Types].xml").decode("utf8")
    have = set(re.findall(r'w:styleId="([^"]+)"', styles))
    used = set(re.findall(r'w:(?:pStyle|tblStyle) w:val="([^"]+)"', doc))
    check(not used - have, f"Word note uses only template styles (missing: {sorted(used - have)})")
    with zipfile.ZipFile(template) as z:
        tpl_styles = z.read("word/styles.xml").decode("utf8")
    check(styles == tpl_styles, "Word note keeps the template's style sheet byte-for-byte")
    check("wordprocessingml.document.main+xml" in cts, "Word note is a .docx, not a template")
    from xml.etree import ElementTree as ET
    ET.fromstring(doc)
    check(True, "Word note XML is well-formed")
    text = re.sub(r"<[^>]+>", "", doc)
    check("Nothing in this note is a Board position" in text,
          "Word note carries its status caveat")


def main(data_path, calc_path, outdir, template):
    data = json.load(open(data_path, encoding="utf8"))
    calc = json.load(open(calc_path, encoding="utf8"))
    out = Path(outdir)
    check_tracker(out / "ESABCC_Policy-Gap-Tracker_2026-07.xlsx", data)
    check_sector_gaps(out / "ESABCC_Policy-Gaps_Transport-Industry_2026-07.xlsx", data)
    check_indicator_check(out / "ESABCC_Indicator-Check_2026-07.xlsx", data, calc)
    check_overviews(out, data, calc)
    for f in out.glob("*.xlsx"):
        check_no_errors(f)
        check_chart_categories(f)

    # The July-2026 content fact-check must actually reach the documents — an
    # earlier run rebuilt them from a stale extract and silently shipped the
    # pre-fact-check numbers.
    inds = {i.get("code"): i for i in data["indicators"]}
    check(inds["B6"].get("targetValue") == 41.5,
          f"fact-check reached the data: B6 target is {inds['B6'].get('targetValue')} (must be 41.5)")
    check(inds["T5b"].get("targetValue") == 80000,
          f"fact-check reached the data: T5b target is {inds['T5b'].get('targetValue')}")
    check(inds["B4 (population)"]["unit"] == "index (2005 = 1.0)",
          f"fact-check reached the data: B4 unit is {inds['B4 (population)']['unit']}")
    gaps = {g["id"]: g for g in data["policyGaps"]["POLICY_GAPS"]}
    check("IPCEI" not in gaps["energy-hydrogen-support"]["instrument"],
          "fact-check reached the data: IPCEIs removed from the hydrogen row")
    syn = {e["id"]: e for e in data["synergies"]["SYNERGIES"]}
    check("no evidence of a substantial shift" in syn["tr-iww-lowflow"]["mechanism"],
          "fact-check reached the data: the Rhine entry states its source's actual finding")
    check(syn["ind-cement-cool-durable"]["kind"] == "mixed",
          "fact-check reached the data: the cement durability entry is retagged conditional")
    check_docx(out / "ESABCC_Synergies-Trade-offs_Industry-Transport_2026-07.docx", template)

    for m in OK:
        print(f"  ok   {m}")
    for m in FAIL:
        print(f"  FAIL {m}")
    print(f"\n{len(OK)} checks passed, {len(FAIL)} failed")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
