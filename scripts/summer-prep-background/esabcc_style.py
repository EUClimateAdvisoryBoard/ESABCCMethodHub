"""
Shared ESABCC look-and-feel for the Summer Prep background workbooks.

The palette and fonts are lifted from the Board's blank Word template
(`Advisory_Board_blank_template.dotx`) so the Excel background documents sit
next to the Word notes without a visual break:

  * theme colour "text2" / dk2  -> #4D7E83  (the ESABCC teal used for headings,
    table header rows and captions)
  * the Data-table style fill   -> #D8E7E8  (teal at 33 % tint, table body)
  * theme accents 1-6           -> #648ECA #B04545 #F59E2D #EFE973 #B487C7
                                   #8EC26A (categorical coding)
  * fonts                       -> Segoe UI / Segoe UI Semibold / Semilight

Everything a workbook needs for a consistent sheet is here: the title block,
the header row, the body-row banding, and the column-width/wrap handling.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette (ESABCC template) ────────────────────────────────────────────────
TEAL = "4D7E83"          # dk2 / text2 — headings, header rows
TEAL_DARK = "3A6165"
TEAL_MID = "8FB4B7"
TEAL_LIGHT = "D8E7E8"    # data-table body fill (text2, 33 % tint)
TEAL_PALE = "EFF6F6"     # banding
INK = "2C2D2D"           # body text
MUTED = "6E7A7B"
RULE = "B9CFD1"
WHITE = "FFFFFF"

ACCENT_BLUE = "648ECA"
ACCENT_RED = "B04545"
ACCENT_ORANGE = "F59E2D"
ACCENT_YELLOW = "EFE973"
ACCENT_PURPLE = "B487C7"
ACCENT_GREEN = "8EC26A"
ORANGE_REC = "EC621D"    # the template's "Recommendation" style colour

FONT = "Segoe UI"
FONT_SEMI = "Segoe UI Semibold"
FONT_LIGHT = "Segoe UI Semilight"

# Gap types and live statuses, mapped onto the template's accents so the
# workbooks and the Word notes read as one family.
GAP_TYPE_COLOR = {
    "policy": ACCENT_RED,
    "ambition": ACCENT_ORANGE,
    "implementation": ACCENT_BLUE,
    "inconsistency": ACCENT_PURPLE,
}
GAP_STATUS_COLOR = {
    "open": ACCENT_RED,
    "partially-addressed": ACCENT_ORANGE,
    "addressed": TEAL,
    "unknown": MUTED,
}
KIND_COLOR = {
    "synergy": TEAL,
    "trade-off": ACCENT_RED,
    "mixed": ACCENT_ORANGE,
}

# ── Reusable part styles ─────────────────────────────────────────────────────
TITLE_FONT = Font(name=FONT_SEMI, size=18, color=TEAL, bold=True)
SUBTITLE_FONT = Font(name=FONT_LIGHT, size=11, color=TEAL)
H2_FONT = Font(name=FONT_SEMI, size=12, color=TEAL, bold=True)
LABEL_FONT = Font(name=FONT_SEMI, size=9, color=TEAL)
BODY_FONT = Font(name=FONT, size=9, color=INK)
SMALL_FONT = Font(name=FONT_LIGHT, size=8, color=MUTED)
NOTE_FONT = Font(name=FONT_LIGHT, size=8.5, color=TEAL)
HEADER_FONT = Font(name=FONT_SEMI, size=9, color=WHITE)
LINK_FONT = Font(name=FONT, size=9, color="0563C1", underline="single")

HEADER_FILL = PatternFill("solid", fgColor=TEAL)
BODY_FILL = PatternFill("solid", fgColor=TEAL_LIGHT)
BAND_FILL = PatternFill("solid", fgColor=TEAL_PALE)
NONE_FILL = PatternFill(fill_type=None)

TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_LEFT = Alignment(horizontal="left", vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

_thin_white = Side(style="thin", color=WHITE)
ROW_BORDER = Border(bottom=_thin_white, top=_thin_white, left=_thin_white, right=_thin_white)


def title_block(ws, title, subtitle=None, note=None, width=8, row=1):
    """Write the standard teal title block; returns the next free row."""
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    ws.row_dimensions[row].height = 26
    r = row + 1
    if subtitle:
        c = ws.cell(row=r, column=1, value=subtitle)
        c.font = SUBTITLE_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        ws.row_dimensions[r].height = 30
        r += 1
    if note:
        c = ws.cell(row=r, column=1, value=note)
        c.font = SMALL_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        r += 1
    return r + 1


def header_row(ws, row, headers, height=30):
    """Write a teal header row across `headers` starting at column 1."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = ROW_BORDER
    ws.row_dimensions[row].height = height


def body_row(ws, row, values, band=False, height=None, wrap_cols=()):
    """Write one banded body row. `wrap_cols` are 1-based column indices."""
    fill = BODY_FILL if band else BAND_FILL
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY_FONT
        c.fill = fill
        c.border = ROW_BORDER
        c.alignment = TOP_WRAP if (not wrap_cols or i in wrap_cols) else TOP_LEFT
    if height:
        ws.row_dimensions[row].height = height


def set_widths(ws, widths, start=1):
    for i, w in enumerate(widths, start=start):
        ws.column_dimensions[get_column_letter(i)].width = w


def tint(ws, cell, color, bold=True):
    """Colour a single cell's text (used for type/status tags)."""
    ws[cell].font = Font(name=FONT_SEMI, size=9, color=color, bold=bold)


def sheet_setup(ws, tab_color=TEAL, freeze=None, zoom=100):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    if freeze:
        ws.freeze_panes = freeze


def note_line(ws, row, text, width=8, font=None):
    c = ws.cell(row=row, column=1, value=text)
    c.font = font or NOTE_FONT
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    return row + 1


def text_categories(chart, ws, col, first_row, last_row):
    """
    Point a chart's category axis at a column of TEXT labels.

    `chart.set_categories()` always writes the reference as a numeric one
    (`c:numRef`). Excel reads that as "these cells are numbers", finds text,
    and draws the axis with no labels at all — the chart comes out as bars
    with nothing naming them. Rewriting the reference as `c:strRef` is what
    makes the names appear.

    Numeric categories (a column of years) are fine through `set_categories`;
    this is only for label columns.
    """
    from openpyxl.chart.data_source import AxDataSource, StrRef

    letter = get_column_letter(col)
    ref = f"'{ws.title}'!${letter}${first_row}:${letter}${last_row}"
    for s in chart.series:
        s.cat = AxDataSource(strRef=StrRef(f=ref))
    return chart


def style_chart(chart, title, height=7.5, width=17):
    """Common chart chrome: ESABCC teal title, no border, light gridlines."""
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (
        CharacterProperties, Font as DrawingFont, Paragraph, ParagraphProperties,
        RichTextProperties,
    )
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.shapes import GraphicalProperties

    chart.title = title
    chart.height = height
    chart.width = width
    chart.style = None
    # Without this the title is drawn ON TOP of the chart instead of above it
    # — see fix_title_overlays() for why, and for the axis titles that are set
    # by the caller after this function has run.
    chart.title.overlay = False
    cp = CharacterProperties(latin=DrawingFont(typeface=FONT_SEMI), sz=1100, b=False,
                             solidFill=TEAL)
    chart.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=cp)
    for p in chart.title.tx.rich.p:
        for r in p.r or []:
            r.rPr = cp
    axis_cp = CharacterProperties(latin=DrawingFont(typeface=FONT), sz=800, solidFill=INK)
    for ax in (chart.x_axis, chart.y_axis):
        ax.txPr = RichText(bodyPr=RichTextProperties(),
                           p=[Paragraph(pPr=ParagraphProperties(defRPr=axis_cp), endParaRPr=axis_cp)])
        ax.majorGridlines = None
        ax.spPr = GraphicalProperties(ln=LineProperties(solidFill=RULE))
        # openpyxl leaves both of these out, and Excel then draws the axis
        # without its tick labels — the chart appears with bars but no names.
        ax.delete = False
        ax.tickLblPos = "nextTo"
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    if chart.legend is not None:
        chart.legend.overlay = False
        chart.legend.txPr = RichText(
            bodyPr=RichTextProperties(),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=axis_cp), endParaRPr=axis_cp)],
        )
    return chart


#: A default row is 15 pt ≈ 0.53 cm. A chart is anchored to a cell but drawn
#: at its own size, floating over the grid, so the NEXT chart in the same
#: column has to be anchored far enough down to clear it — the sheet's own row
#: count says nothing about that.
ROW_CM = 0.53


def rows_for(height_cm, clearance=1):
    """Rows a chart of `height_cm` covers, plus a row or two of breathing room."""
    return int(height_cm / ROW_CM) + 1 + clearance


def fix_title_overlays(wb):
    """
    Stop Excel drawing chart and axis titles on top of the labels.

    openpyxl writes a `<c:title>` with no `<c:overlay>` child. The element is
    optional, but Excel's reading of an absent one is "overlay", not "don't":
    it lays the title over the plot instead of reserving space above it. The
    visible result is a chart title sitting across the top category labels and
    — worse, because it happens on every chart with a unit — a rotated y-axis
    title printed straight through the tick numbers.

    `style_chart` sets the flag on the chart title, but axis titles are
    assigned by the callers afterwards (`ch.y_axis.title = ind["unit"]`), each
    assignment building a fresh Title object with the flag unset again. Rather
    than chase every call site, this runs over the finished workbook and fixes
    every title it can reach. Call it immediately before `wb.save()`.

    Returns the number of titles corrected.
    """
    fixed = 0
    for ws in wb.worksheets:
        for chart in getattr(ws, "_charts", []):
            for holder in (chart, chart.x_axis, chart.y_axis):
                title = getattr(holder, "title", None)
                if title is not None and getattr(title, "overlay", None) is not False:
                    title.overlay = False
                    fixed += 1
    return fixed
