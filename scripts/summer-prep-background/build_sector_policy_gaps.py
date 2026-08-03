"""
Background document 3 — Policy gaps for TRANSPORT and INDUSTRY.

Renders the Summer Prep note that extends the Policy Gap Tracker
(`src/data/summer-prep-sector-gaps.ts`, joined to `src/data/policy-gaps.ts`):

  1. the transport/industry gaps of the 2024 report, re-assessed against
     legislation adopted since (provisional mid-2026 read);
  2. candidate additional gaps for the sector lead to accept or reject;
  3. the gap landscape per sector and subsector.

Usage:  python3 build_sector_policy_gaps.py data.json out.xlsx
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from esabcc_style import (
    ACCENT_PURPLE, BODY_FILL, BODY_FONT, FONT, FONT_SEMI, GAP_STATUS_COLOR, GAP_TYPE_COLOR,
    H2_FONT, INK, LABEL_FONT, MUTED, SMALL_FONT, TEAL, TEAL_PALE, WHITE, body_row,
    fix_title_overlays, header_row, note_line, rows_for, set_widths, sheet_setup, style_chart,
    text_categories, title_block,
)

PREPARED = "27 July 2026"
LONG = "Landscape data"


def sources_cell(urls):
    return "\n".join(urls or []) or "—"


def build(data, out_path):
    sg = data["sectorGaps"]
    pg = data["policyGaps"]
    gaps_by_id = {g["id"]: g for g in pg["POLICY_GAPS"]}
    type_meta = pg["GAP_TYPE_META"]
    status_meta = pg["GAP_STATUS_META"]
    meta = sg["SECTOR_GAP_META"]
    subsectors = sg["SECTOR_SUBSECTORS"]
    reassessments = sg["GAP_REASSESSMENTS"]
    candidates = sg["CANDIDATE_GAPS"]

    order = {"Industry": 0, "Transport": 1}
    reassessed = sorted(
        ((gid, r, gaps_by_id[gid]) for gid, r in reassessments.items() if gid in gaps_by_id),
        key=lambda x: (order.get(x[2]["sector"], 9),
                       subsectors.get(x[2]["sector"], []).index(x[1]["subsector"])
                       if x[1]["subsector"] in subsectors.get(x[2]["sector"], []) else 9),
    )
    cands = sorted(candidates, key=lambda c: (order.get(c["sector"], 9),
                                              subsectors.get(c["sector"], []).index(c["subsector"])
                                              if c["subsector"] in subsectors.get(c["sector"], []) else 9))

    wb = Workbook()

    # ── Read me ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Read me"
    sheet_setup(ws)
    set_widths(ws, [24, 24, 24, 24, 24, 20, 16, 16])
    r = title_block(
        ws,
        "Policy gaps — Transport & Industry",
        "The transport- and industry-tagged findings of the 2024 ESABCC progress report, re-assessed "
        "against the legislation adopted since; plus candidate additional gaps and the per-subsector "
        "gap landscape.",
        f"Background document for the Summer Prep · Policy Gap 2.0 Report module (Note 3) · prepared {PREPARED}. "
        f"Re-assessment vintage: {meta['reassessmentAsOf']}.",
    )

    ws.cell(row=r, column=1, value="What is in this workbook").font = H2_FONT
    r += 1
    for i, (name, desc) in enumerate([
        ("Re-assessed gaps", f"The {len(reassessed)} transport/industry gaps the report tagged, each with its "
                             "verbatim report quote and page, its assigned subsector, and a provisional mid-2026 "
                             "read of whether it still exists — with the sources behind that read."),
        ("Candidate gaps", f"{len(cands)} issues that look like policy, ambition or implementation gaps but are "
                           "not (yet) tagged as such in the report. Each carries the test that would confirm or "
                           "refute it. These are proposals for the sector lead, not findings."),
        ("Gap landscape", "Gaps per sector and subsector — report gaps by live status, and candidates — with the "
                          "landscape figure. Counts are COUNTIFS formulas over the “Landscape data” sheet."),
        (LONG, "The same gaps in long format, one row per gap × subsector. A gap that bears on two subsectors "
               "appears once per subsector, so no subsector row is silently empty. This is the sheet the "
               "landscape counts read from."),
        ("Legend", "Gap types, live statuses and the subsector taxonomy."),
    ]):
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT_SEMI, size=9, color=TEAL)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
        c = ws.cell(row=r, column=2, value=desc)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = BODY_FILL if i % 2 == 0 else PatternFill("solid", fgColor=TEAL_PALE)
        ws.row_dimensions[r].height = 44
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="How to read it — and what it is not").font = H2_FONT
    r += 1
    for line in [
        "• The base facts are the Board's: that the report tagged the gap, its verbatim quote, and its page. Those "
        "columns are unchanged from the report and should not be edited.",
        "• The re-assessment status and note, the candidate gaps, and the assignment of each gap to a subsector are "
        "AI-compiled working material for this prep cycle and are PENDING VERIFICATION. Nothing here is a Board "
        "position — it is a prompt for the sector lead's own judgement.",
        "• Where a re-assessment leans on something that happened after the report, the source URLs are in the "
        "“Sources” column. Check those first when you take a row forward.",
        "• Candidate gaps are deliberately falsifiable: each has a “Test to confirm / refute”. Work the test, then "
        "accept the candidate into the tracker or drop it — and record which.",
        "• The live version is in the Method Hub (Summer Prep › Policy Gap 2.0 Report › Policy Gaps — Transport & "
        "Industry).",
    ]:
        r = note_line(ws, r, line, width=8, font=BODY_FONT)
        ws.row_dimensions[r - 1].height = 28
    r += 1

    ws.cell(row=r, column=1, value="Source").font = H2_FONT
    r += 1
    for label, value in [
        ("Report", f"European Scientific Advisory Board on Climate Change (2024) “{meta['reportTitle']}”, "
                   f"{meta['reportPublished']}."),
        ("Report URL", meta["reportUrl"]),
        ("Data source", "src/data/summer-prep-sector-gaps.ts, joined to src/data/policy-gaps.ts, in the ESABCC "
                        "Method Hub repository."),
        ("Vintage", meta["reassessmentAsOf"]),
    ]:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
        r += 1

    # ── Re-assessed gaps ─────────────────────────────────────────────────────
    ra = wb.create_sheet("Re-assessed gaps")
    sheet_setup(ra, freeze="D2")
    header_row(ra, 1, [
        "#", "Sector", "Subsector", "Gap type", "Finding", "What the Board found",
        "Main EU instrument(s)", "Verbatim quote from the report", "Report reference",
        "Status mid-2026 (provisional)", "Re-assessment — what has changed since January 2024",
        "Also bears on", "Sources for the re-assessment", "ID",
    ], height=34)
    set_widths(ra, [4, 11, 16, 15, 28, 40, 24, 52, 20, 18, 56, 16, 34, 24])
    for i, (gid, rs, g) in enumerate(reassessed):
        row = i + 2
        body_row(ra, row, [
            i + 1, g["sector"], rs["subsector"], type_meta[g["type"]]["label"], g["title"],
            g["description"], g["instrument"], "“" + g["quote"] + "”", g["reference"],
            status_meta[rs["status"]]["label"], rs["note"],
            ", ".join(rs.get("alsoSubsectors") or []) or "—",
            sources_cell(rs.get("sources")), gid,
        ], band=(i % 2 == 0))
        ra.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ra.cell(row=row, column=4).font = Font(name=FONT_SEMI, size=9, color=GAP_TYPE_COLOR[g["type"]])
        ra.cell(row=row, column=5).font = Font(name=FONT_SEMI, size=9, color=INK)
        ra.cell(row=row, column=8).font = Font(name="Segoe UI Semilight", size=8.5, color=MUTED, italic=True)
        ra.cell(row=row, column=10).font = Font(name=FONT_SEMI, size=9, color=GAP_STATUS_COLOR[rs["status"]])
        ra.cell(row=row, column=13).font = Font(name=FONT, size=7.5, color="0563C1")
        ra.cell(row=row, column=14).font = SMALL_FONT
        ra.row_dimensions[row].height = 92
    ra.auto_filter.ref = f"A1:N{len(reassessed) + 1}"
    note_line(ra, len(reassessed) + 3,
              "Notes: columns A–I are the report's own findings (unchanged). Columns J–M are a provisional mid-2026 "
              "re-assessment compiled for this prep cycle and pending verification — not a Board position.", width=11)
    note_line(ra, len(reassessed) + 4,
              f"Source: ESABCC (2024) “{meta['reportTitle']}”, {meta['reportPublished']}; re-assessment "
              f"{meta['reassessmentAsOf']}.", width=11)

    # ── Candidate gaps ───────────────────────────────────────────────────────
    cg = wb.create_sheet("Candidate gaps")
    sheet_setup(cg, freeze="D2")
    header_row(cg, 1, [
        "#", "Sector", "Subsector", "Proposed gap type", "Candidate gap",
        "Why this may be a gap", "Instrument(s) concerned",
        "Test to confirm / refute", "Sources", "Lead's decision", "ID",
    ], height=34)
    set_widths(cg, [4, 11, 18, 16, 30, 56, 26, 56, 32, 16, 24])
    for i, c in enumerate(cands):
        row = i + 2
        body_row(cg, row, [
            i + 1, c["sector"], c["subsector"], type_meta[c["type"]]["label"], c["title"],
            c["rationale"], c["instrument"], c["testToConfirm"], sources_cell(c.get("sources")),
            "", c["id"],
        ], band=(i % 2 == 0))
        cg.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
        cg.cell(row=row, column=4).font = Font(name=FONT_SEMI, size=9, color=GAP_TYPE_COLOR[c["type"]])
        cg.cell(row=row, column=5).font = Font(name=FONT_SEMI, size=9, color=INK)
        cg.cell(row=row, column=9).font = Font(name=FONT, size=7.5, color="0563C1")
        cg.cell(row=row, column=10).fill = PatternFill("solid", fgColor="FFF6D9")
        cg.cell(row=row, column=11).font = SMALL_FONT
        cg.row_dimensions[row].height = 96
    cg.auto_filter.ref = f"A1:K{len(cands) + 1}"
    note_line(cg, len(cands) + 3,
              "Notes: “Lead's decision” (shaded) is the column to fill in — accept into the gap tracker, reject, or "
              "park with a reason. Every candidate is AI-compiled working material pending verification; none is a "
              "Board finding.", width=9)

    # ── Landscape data (long format) ─────────────────────────────────────────
    ld = wb.create_sheet(LONG)
    sheet_setup(ld, freeze="A2")
    header_row(ld, 1, ["Sector", "Subsector", "Origin", "Live status", "Gap type", "Finding", "ID"], height=20)
    set_widths(ld, [12, 32, 14, 20, 18, 52, 24])
    rows = []
    for gid, rs, g in reassessed:
        for sub in [rs["subsector"]] + list(rs.get("alsoSubsectors") or []):
            rows.append([g["sector"], sub, "Report gap", status_meta[rs["status"]]["label"],
                         type_meta[g["type"]]["label"], g["title"], gid])
    for c in cands:
        rows.append([c["sector"], c["subsector"], "Candidate", "Candidate — to test",
                     type_meta[c["type"]]["label"], c["title"], c["id"]])
    for i, vals in enumerate(rows):
        body_row(ld, i + 2, vals, band=(i % 2 == 0), height=16)
        ld.cell(row=i + 2, column=3).font = Font(
            name=FONT_SEMI, size=9, color=TEAL if vals[2] == "Report gap" else ACCENT_PURPLE)
    ld.auto_filter.ref = f"A1:G{len(rows) + 1}"
    note_line(ld, len(rows) + 3,
              "One row per gap × subsector: a gap flagged as bearing on a second subsector appears twice, so the "
              "landscape counts do not leave that subsector empty. Row totals here are therefore larger than the "
              "number of distinct gaps.", width=7)

    # ── Gap landscape ────────────────────────────────────────────────────────
    gl = wb.create_sheet("Gap landscape")
    sheet_setup(gl)
    set_widths(gl, [34, 12, 16, 12, 14, 12, 3, 20, 12])
    r = title_block(
        gl,
        "The gap landscape, subsector by subsector",
        "How the transport and industry gaps distribute across the report-aligned subsectors: report gaps by their "
        "provisional mid-2026 status, and the candidate gaps proposed on top. Counts are COUNTIFS formulas over the "
        "“Landscape data” sheet.",
        width=9,
    )
    status_cols = [status_meta[s]["label"] for s in ("open", "partially-addressed", "addressed")]
    cols = status_cols + ["Candidate — to test"]
    chart_anchor = r
    matrix_rows = []
    for sector in ("Industry", "Transport"):
        gl.cell(row=r, column=1, value=sector).font = H2_FONT
        r += 1
        header_row(gl, r, ["Subsector"] + cols + ["Total"], height=32)
        head = r
        first = r + 1
        for i, sub in enumerate(subsectors[sector]):
            row = first + i
            vals = [sub]
            for j, label in enumerate(cols):
                col = get_column_letter(2 + j)
                vals.append(
                    f'=COUNTIFS(\'{LONG}\'!$A$2:$A$200,"{sector}",'
                    f'\'{LONG}\'!$B$2:$B$200,$A{row},\'{LONG}\'!$D$2:$D$200,{col}${head})'
                )
            vals.append(f"=SUM(B{row}:E{row})")
            body_row(gl, row, vals, band=(i % 2 == 0), height=17)
            gl.cell(row=row, column=1).font = Font(name=FONT_SEMI, size=9, color=INK)
            for col in range(2, 7):
                gl.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            matrix_rows.append(row)
        last = first + len(subsectors[sector]) - 1
        trow = last + 1
        body_row(gl, trow, [f"All {sector.lower()}"] + [
            f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{last})" for c in range(2, 7)
        ], band=True, height=18)
        for col in range(1, 7):
            c = gl.cell(row=trow, column=col)
            c.font = Font(name=FONT_SEMI, size=9, color=WHITE)
            c.fill = PatternFill("solid", fgColor=TEAL)
            if col > 1:
                c.alignment = Alignment(horizontal="center", vertical="center")

        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.add_data(Reference(gl, min_col=2, max_col=5, min_row=head, max_row=last),
                       titles_from_data=True)
        chart.set_categories(Reference(gl, min_col=1, min_row=first, max_row=last))
        style_chart(chart, f"{sector} — gaps per subsector", height=7.4, width=15)
        text_categories(chart, gl, 1, first, last)
        chart.x_axis.title = "Number of gaps"
        colors = [GAP_STATUS_COLOR["open"], GAP_STATUS_COLOR["partially-addressed"],
                  GAP_STATUS_COLOR["addressed"], ACCENT_PURPLE]
        for series, color in zip(chart.series, colors):
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.noFill = True
        chart.legend.position = "b"
        gl.add_chart(chart, f"H{chart_anchor}")
        chart_anchor = max(trow + 2, chart_anchor + rows_for(7.4))
        r = trow + 3

    note_line(gl, r, "Notes: a gap that bears on more than one subsector is counted in each — the totals are "
                     "therefore assignments, not distinct gaps. “Candidate — to test” rows are proposals, not Board "
                     "findings.", width=6)
    note_line(gl, r + 1, f"Source: ESABCC (2024) “{meta['reportTitle']}”; re-assessment {meta['reassessmentAsOf']}.",
              width=6)

    # ── Legend ───────────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    sheet_setup(lg)
    set_widths(lg, [26, 76, 12])
    r = title_block(lg, "Legend", "The report's gap classification, the status vocabulary used in the "
                                 "re-assessment, and the subsector taxonomy.", width=3)
    header_row(lg, r, ["Gap type", "What it means (report methodology)", "Colour"], height=20)
    r += 1
    for i, (key, m) in enumerate(type_meta.items()):
        body_row(lg, r, [m["label"], m["description"], ""], band=(i % 2 == 0), height=20)
        lg.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=GAP_TYPE_COLOR[key])
        lg.cell(row=r, column=3).fill = PatternFill("solid", fgColor=GAP_TYPE_COLOR[key])
        r += 1
    r += 2
    header_row(lg, r, ["Status (mid-2026 re-assessment)", "What it means here", "Colour"], height=20)
    r += 1
    help_ = {
        "open": "Nothing adopted since the report changes the finding — assessed unchanged.",
        "partially-addressed": "Legislation since the report narrows the gap but does not close it.",
        "addressed": "An adopted instrument closes the gap; any residual risk is named in the note.",
        "unknown": "Not yet re-assessed.",
    }
    for i, (key, m) in enumerate(status_meta.items()):
        body_row(lg, r, [m["label"], help_[key], ""], band=(i % 2 == 0), height=20)
        lg.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=GAP_STATUS_COLOR[key])
        lg.cell(row=r, column=3).fill = PatternFill("solid", fgColor=GAP_STATUS_COLOR[key])
        r += 1
    body_row(lg, r, ["Candidate — to test", "Proposed additional gap, not tagged in the report. Work the test in the "
                                            "“Candidate gaps” sheet, then accept or reject.", ""], band=True, height=28)
    lg.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=ACCENT_PURPLE)
    lg.cell(row=r, column=3).fill = PatternFill("solid", fgColor=ACCENT_PURPLE)
    r += 3
    header_row(lg, r, ["Sector", "Subsectors (report-aligned taxonomy)", ""], height=20)
    r += 1
    for i, (sector, subs) in enumerate(subsectors.items()):
        body_row(lg, r, [sector, " · ".join(subs), ""], band=(i % 2 == 0), height=20)
        lg.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=TEAL)
        r += 1

    # No cached results are written for the COUNTIFS formulas, so ask the
    # spreadsheet app to calculate the whole book the moment it opens.
    wb.calculation.fullCalcOnLoad = True
    wb.properties.title = "Policy gaps — Transport & Industry (Summer Prep note 3)"
    wb.properties.subject = "Summer Prep · Policy Gap 2.0 Report — background document"
    wb.properties.creator = "ESABCC Method Hub"
    fix_title_overlays(wb)
    wb.save(out_path)
    print(f"wrote {out_path}: {len(reassessed)} re-assessed, {len(cands)} candidates, "
          f"{len(rows)} landscape assignments")


if __name__ == "__main__":
    build(json.load(open(sys.argv[1], encoding="utf8")), sys.argv[2])
