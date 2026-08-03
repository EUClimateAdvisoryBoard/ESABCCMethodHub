"""
Combined background document — one workbook for the whole Summer Prep
Indicator module.

Merges the three separate indicator workbooks into a single file:

  * Indicator Check   — Overview, one tab per chapter, Data (long),
                        Where the data comes from  (build_indicator_check.py)
  * New data overview — the one "everything that has moved" figure
                        (build_indicator_overview.build_new_data_overview)
  * Old vs new         — Old vs new, Derivations, Sources
                        (build_indicator_overview.build_old_vs_new)

Each of those builders already knows how to write into a caller-supplied
Workbook instead of creating and saving its own (pass `wb=`); this script is
the caller. Every source builder's own "Read me" sheet is skipped in that
mode — this script writes ONE merged "Read me" that describes every section
and says which sheets belong to it, so the combined book still opens on an
orientation page instead of straight into a table.

The exact-dataset-link column ("Where the data comes from", "Sources") and
the Primary-source hyperlinks ("Overview", "Old vs new") are added inside the
source builders themselves (see dataset_links.py) — they appear here and in
the three standalone workbooks alike, since it is the same code either way.

Usage:
  python3 build_indicator_combined.py data.json calc.json calc_excel.json factcheck.json reportway.json out.xlsx

  data.json      — extract.mjs's output (src/data, run live, no cache)
  calc.json      — the calc grid, "Indicator Check" shape (columns[].formula
                   as a string, rows[].cells) — used for "Where the data
                   comes from" / the chapter tabs' method label
  calc_excel.json — the calc grid, "Old vs new" / Derivations shape
                   (columns[].expr, rows[].formulas as ready Excel formula
                   text) — used for the Derivations sheet
  factcheck.json — the 27 July 2026 fact-check results (rows[])
  reportway.json — the report's-own-derivation-continued rows ({"filled": {…}})
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import build_dashboard as dashboard
import build_indicator_check as check
import build_indicator_overview as overview
from link_derivations import DerivationLinks
from esabcc_style import (
    BODY_FILL, BODY_FONT, FONT_SEMI, H2_FONT, TEAL, TEAL_PALE,
    header_row, note_line, set_widths, sheet_setup, title_block,
)

PREPARED = "27 July 2026"

FC_ORDER = ["CONFIRMED", "NOT CHECKABLE", "NO SOURCE YEAR", "REVISION", "WRONG"]


def _collapse_factcheck(fcraw):
    """Per indicator: the worst verdict over its points (build_indicator_overview's own rule)."""
    fc = {}
    for row in fcraw.get("rows", []):
        cur = fc.setdefault(row["id"], {"verdict": "CONFIRMED"})
        if FC_ORDER.index(row["verdict"]) > FC_ORDER.index(cur["verdict"]):
            cur["verdict"] = row["verdict"]
    return fc


def build(data, calc, calc_excel, factcheck, reportway, out_path=None, wb=None):
    """
    Render the combined workbook. Mirrors the wb=None / wb=<Workbook> contract
    of the three source builders: standalone when `wb` is None (creates,
    writes the merged Read me, saves to `out_path`); embeddable otherwise.
    """
    standalone = wb is None
    inds = data["indicators"]
    fc = _collapse_factcheck(factcheck)
    reads = {i["id"]: overview.read(i) for i in inds}

    if standalone:
        wb = Workbook()
        wb.remove(wb.active)

    # ── the three source builders, into the same workbook ───────────────────
    # They run BEFORE the Read me because the Read me describes what the
    # Derivations wiring managed to connect, which is only known once the
    # Derivations sheet exists.
    links = DerivationLinks()
    check.build(data, calc, factcheck, wb=wb, links=links)
    overview.build_new_data_overview(inds, reads, fc, wb=wb, title="New data", links=links)
    overview.build_old_vs_new(inds, reads, calc_excel, fc, reportway, wb=wb, links=links)

    backed = links.backed_indicators([i["id"] for i in inds])
    dashboard.build_dashboard(inds, reads, factcheck, backed, wb, PREPARED, links=links)
    links.apply(wb)

    # ── merged Read me ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Read me", 0)
    sheet_setup(ws)
    set_widths(ws, [24, 22, 22, 22, 22, 20, 16, 16])
    updated = [i for i in inds if reads[i["id"]]["post"]]
    r = title_block(
        ws,
        "ESABCC Indicators — the combined background workbook",
        f"Every progress indicator of the 2024 ESABCC report in one file: the indicator check (what has moved "
        f"since the report, by chapter), the one figure carrying every move at once, and the report's own "
        f"figures next to the database's latest values with the arithmetic between them written out as a live "
        f"Excel formula. {len(updated)} of the {len(inds)} indicators have gained data since January 2024.",
        f"Background document for the Summer Prep · Policy Gap 2.0 Report module · prepared {PREPARED}. "
        "This merges the three separate indicator workbooks into one — nothing on any sheet has changed by "
        "being combined.",
    )

    ws.cell(row=r, column=1, value="What is in this workbook").font = H2_FONT
    r += 1
    groups = [
        ("Dashboard", "The module at a glance: key figures, the largest moves since the report as a diverging "
                     "bar, chapter coverage, four headline trend lines and — where a target exists — a "
                     f"distance-to-target panel. Its source blocks live on the hidden '{dashboard.DATA_SHEET}' "
                     "sheet."),
        ("Overview", "All indicators in one table — baseline, latest value, change — plus the two overview "
                     "figures: the largest moves since the report, and how many indicators have new data per "
                     "chapter. Part of the Indicator Check."),
        ("Emissions … Fairness, Adaptation (one tab per chapter)", "The chapter's indicators as a table, then "
                     "every indicator's full series as its own small table and line chart. Part of the "
                     "Indicator Check."),
        ("Data (long)", "Every data point of every indicator in one flat table — the sheet to pivot or filter. "
                     "Part of the Indicator Check."),
        ("Where the data comes from", "For every point added since the report: the calc inputs it is built "
                     "from, the formula, the exact source query — with a DOI link where the dataset has one — "
                     "and the 27 July 2026 fact-check verdict. Part of the Indicator Check."),
        ("New data", "Everything that has moved since the report in one figure, indexed to each indicator's own "
                     "report baseline so indicators in wildly different units are still comparable."),
        ("Old vs new", "One row per indicator: the report's own figure, the database's latest value, the "
                     "change, and how that later figure was obtained — with a Primary source hyperlink where "
                     "the dataset has a DOI."),
        ("Derivations", "The calc grid behind every indicator as LIVE EXCEL FORMULAS — click a Value cell and "
                     "Excel shows the arithmetic over the inputs in the same block."),
        ("Sources", "The exact source query behind every input column in Derivations, with a DOI link where "
                     "the dataset has one."),
    ]
    header_row(ws, r, ["Sheet group", "What it holds", "", "", "", "", "", ""], height=20)
    r += 1
    for i, (name, desc) in enumerate(groups):
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT_SEMI, size=9, color=TEAL)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        c = ws.cell(row=r, column=2, value=desc)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = BODY_FILL if i % 2 == 0 else PatternFill("solid", fgColor=TEAL_PALE)
        ws.row_dimensions[r].height = 46
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="How the numbers stay in sync").font = H2_FONT
    r += 1
    for line in [
        f"• Derivations is the single source of truth for every derivation-backed indicator ({backed} of the "
        f"{len(inds)}): edit an input cell there, and every other cell built from it — Overview, the chapter "
        "tabs, Data (long), New data, Old vs new, and the Dashboard — recomputes on the workbook's next "
        "recalculation, because those cells are live formulas pointing at Derivations, not pasted numbers.",
        f"• A cell only carries that live link where the Derivations value it points at already agrees with "
        f"the stored data point to within {links.tolerance:.1%} — a bigger gap is left as a plain number "
        "instead of being wired over silently, on the view that a genuine derivation-vs-data discrepancy "
        "should stay visible.",
        f"• Indicators with no Derivations block ({len(inds) - backed} of the {len(inds)}) keep their figures "
        "as plain numbers throughout, exactly as before this workbook wired the rest together.",
    ]:
        r = note_line(ws, r, line, width=8, font=BODY_FONT)
        ws.row_dimensions[r - 1].height = 34
    r += 1

    ws.cell(row=r, column=1, value="Exact dataset links").font = H2_FONT
    r += 1
    for line in [
        "• Where a source names one of the 17 Eurostat datasets behind these indicators, its cell is a live "
        "hyperlink to that dataset's DOI (https://doi.org/10.2908/<CODE>), which resolves to the Eurostat "
        "databrowser page. Non-Eurostat publishers (EEA, EAFO, IRENA, EHPA, SolarPower Europe, WindEurope, "
        "UNFCCC/CRF, Eurofer, Cembureau, BloombergNEF) do not mint DOIs, so their cell links to the canonical "
        "landing page instead.",
        "• A source that names no recognised dataset or publisher is left as plain text — no link is guessed.",
    ]:
        r = note_line(ws, r, line, width=8, font=BODY_FONT)
        ws.row_dimensions[r - 1].height = 34
    r += 1

    ws.cell(row=r, column=1, value="Source").font = H2_FONT
    r += 1
    for label, value in [
        ("Indicators", "The progress indicators of ESABCC (2024) “Towards EU climate neutrality: Progress, "
                       "policy gaps and opportunities”."),
        ("Data source", "The Policy Gap 2.0 Project Workspace indicator database (src/data/esabcc-indicators.ts "
                        "and the calc grids in supabase/migrations/045, 052, 078, 079, 080)."),
        ("Vintage", f"Compiled {PREPARED}. Source publishers revise: re-check before quoting a figure externally."),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_SEMI, size=9, color=TEAL)
        c = ws.cell(row=r, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 34
        r += 1

    # The book opens on the orientation page, then the dashboard; the
    # dashboard's hidden source blocks sit directly behind it.
    for target, name in enumerate(["Read me", "Dashboard", dashboard.DATA_SHEET]):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=target - wb.sheetnames.index(name))

    if standalone:
        wb.calculation.fullCalcOnLoad = True
        wb.properties.title = "ESABCC Indicators — combined background workbook"
        wb.properties.subject = "Summer Prep · Policy Gap 2.0 Report — combined background document"
        wb.properties.creator = "ESABCC Method Hub"
        wb.save(out_path)
        print(f"wrote {out_path}: {len(wb.sheetnames)} sheets, {len(inds)} indicators, "
              f"{len(updated)} with new data, {links.linked} cells wired to Derivations "
              f"({links.held_back} held back as plain numbers)")
    return wb


def main(data_path, calc_path, calc_excel_path, fc_path, reportway_path, out_path):
    data = json.load(open(data_path, encoding="utf8"))
    calc = json.load(open(calc_path, encoding="utf8"))
    calc_excel = json.load(open(calc_excel_path, encoding="utf8"))
    factcheck = json.load(open(fc_path, encoding="utf8"))
    reportway = json.load(open(reportway_path, encoding="utf8"))["filled"]
    build(data, calc, calc_excel, factcheck, reportway, out_path)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6])
