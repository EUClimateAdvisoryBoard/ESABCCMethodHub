"""
Background document 1 — ESABCC Policy Gap Tracker (all 12 report chapters).

Renders `src/data/policy-gaps.ts` (the Method Hub's Policy Gap Tracker module,
Summer Prep › Policy Gap 2.0 Report) as a standalone, branded workbook so the
team can work with the gap list while the Hub owner is away.

Usage:  python3 build_policy_gap_tracker.py data.json out.xlsx
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from esabcc_style import (
    ACCENT_BLUE, ACCENT_ORANGE, ACCENT_PURPLE, ACCENT_RED, BODY_FONT, GAP_STATUS_COLOR,
    GAP_TYPE_COLOR, H2_FONT, LABEL_FONT, MUTED, NOTE_FONT, SMALL_FONT, TEAL, TEAL_LIGHT,
    BODY_FILL, FONT, FONT_SEMI, INK, LINK_FONT, WHITE, body_row, fix_title_overlays, header_row,
    note_line, rows_for, set_widths, sheet_setup, style_chart, text_categories, title_block,
)

PREPARED = "27 July 2026"
TRACKER = "Gap tracker"


def build(data, out_path):
    pg = data["policyGaps"]
    gaps = pg["POLICY_GAPS"]
    type_meta = pg["GAP_TYPE_META"]
    status_meta = pg["GAP_STATUS_META"]
    sectors = pg["GAP_SECTORS"]
    meta = pg["GAP_REPORT_META"]

    # Keep the tracker in the module's own order: sector (report chapter order),
    # then gap type, then title.
    type_order = list(type_meta.keys())
    gaps = sorted(
        gaps,
        key=lambda g: (sectors.index(g["sector"]) if g["sector"] in sectors else 99,
                       type_order.index(g["type"]), g["title"]),
    )

    wb = Workbook()

    # ── 1. Read me ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Read me"
    sheet_setup(ws)
    set_widths(ws, [26, 22, 22, 22, 22, 22, 18, 18])
    r = title_block(
        ws,
        "Policy Gap Tracker — ESABCC 2024 progress report",
        "Every policy, ambition and implementation gap and inconsistency the Board identified "
        "across the 12 chapters of “Towards EU climate neutrality: Progress, policy gaps and "
        "opportunities” (January 2024), with the verbatim report sentence and page behind each finding.",
        f"Background document for the Summer Prep · Policy Gap 2.0 Report module · prepared {PREPARED}.",
    )

    ws.cell(row=r, column=1, value="What is in this workbook").font = H2_FONT
    r += 1
    sheets = [
        (TRACKER, f"All {len(gaps)} findings, one row each: sector/chapter, gap type, the finding, "
                  "the EU instrument(s) concerned, the verbatim report quote, the page reference, "
                  "and two status columns to maintain (see below). Filter with the arrows in row 1."),
        ("Summary", "Counts by chapter × gap type and by live status, with the two overview figures. "
                    "Every number is a COUNTIFS formula over the tracker — it updates by itself when "
                    "you edit, add or delete rows there."),
        ("Legend", "What the four gap types and the four live statuses mean, in the report's own words."),
    ]
    header_row(ws, r, ["Sheet", "What it holds", "", "", "", "", "", ""], height=20)
    r += 1
    for i, (name, desc) in enumerate(sheets):
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT_SEMI, size=9, color=TEAL)
        c = ws.cell(row=r, column=2, value=desc)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = BODY_FILL if i % 2 == 0 else PatternFill("solid", fgColor="EFF6F6")
        ws.row_dimensions[r].height = 42
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="How to use it").font = H2_FONT
    r += 1
    for line in [
        "• The two columns to maintain are “Live status” and “Status note”. Everything to their left is the "
        "January-2024 report baseline and should not be overwritten — if a finding changes, record it in the note.",
        "• “Live status” has a drop-down (Still open / Partially addressed / Addressed / To review). All 66 rows are "
        "seeded to the report baseline “Still open”, so nothing is silently claimed to be solved.",
        "• Rows you add at the bottom are picked up by the Summary sheet automatically (the formulas run to row 400). "
        "Mark them “Custom” in the Origin column so they stay distinguishable from Board findings.",
        "• The same list lives in the Method Hub (Summer Prep › Policy Gap 2.0 Report › Policy Gap Tracker), where it "
        "is filterable and editable in the browser. This workbook is the offline copy — if both are edited, the Hub is "
        "the one to reconcile against.",
    ]:
        r = note_line(ws, r, line, width=8, font=BODY_FONT)
        ws.row_dimensions[r - 1].height = 26
    r += 1

    ws.cell(row=r, column=1, value="Source").font = H2_FONT
    r += 1
    for label, value in [
        ("Report", f"{meta['author']} (2024) “{meta['title']}”, {meta['published']}."),
        ("Report URL", meta["url"]),
        ("Data source", "src/data/policy-gaps.ts in the ESABCC Method Hub repository — every quote was checked "
                        "line-by-line against the report text, and each page reference points at the page carrying "
                        "the tagged finding."),
        ("Status of the content", "The findings and quotes are the Board's own, from the published report. The live "
                                  "status columns are a working tool for the Secretariat, not a Board position."),
    ]:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1

    # ── 2. Gap tracker ───────────────────────────────────────────────────────
    tr = wb.create_sheet(TRACKER)
    sheet_setup(tr, freeze="C2")
    headers = [
        "#", "Sector / report chapter", "Gap type", "Finding", "What the Board found",
        "Main EU instrument(s)", "Verbatim quote from the report", "Report reference",
        "Status in the report (Jan 2024)", "Live status", "Status note (what changed since)",
        "Origin", "ID",
    ]
    header_row(tr, 1, headers, height=34)
    set_widths(tr, [4, 18, 15, 30, 46, 26, 58, 22, 16, 16, 40, 9, 24])

    for i, g in enumerate(gaps):
        row = i + 2
        body_row(tr, row, [
            i + 1,
            g["sector"],
            type_meta[g["type"]]["label"],
            g["title"],
            g["description"],
            g["instrument"],
            "“" + g["quote"] + "”",
            g["reference"],
            "Open",
            status_meta[g["currentStatus"]]["label"],
            g.get("statusNote") or "",
            "Report" if g.get("source", "report") == "report" else "Custom",
            g["id"],
        ], band=(i % 2 == 0))
        tr.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
        tr.cell(row=row, column=3).font = Font(name=FONT_SEMI, size=9,
                                               color=GAP_TYPE_COLOR[g["type"]])
        tr.cell(row=row, column=4).font = Font(name=FONT_SEMI, size=9, color=INK)
        tr.cell(row=row, column=7).font = Font(name="Segoe UI Semilight", size=8.5,
                                               color=MUTED, italic=True)
        tr.cell(row=row, column=10).font = Font(name=FONT_SEMI, size=9,
                                                color=GAP_STATUS_COLOR[g["currentStatus"]])
        tr.cell(row=row, column=13).font = SMALL_FONT
        tr.row_dimensions[row].height = 62

    last = len(gaps) + 1
    tr.auto_filter.ref = f"A1:M{last}"

    # Live-status drop-down, and colour the status cells by their value.
    dv = DataValidation(
        type="list",
        formula1='"Still open,Partially addressed,Addressed,To review"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.prompt = "Pick the live status — the report baseline for every Board finding is “Still open”."
    dv.promptTitle = "Live status"
    tr.add_data_validation(dv)
    dv.add(f"J2:J400")
    for label, color in [
        ("Still open", GAP_STATUS_COLOR["open"]),
        ("Partially addressed", GAP_STATUS_COLOR["partially-addressed"]),
        ("Addressed", GAP_STATUS_COLOR["addressed"]),
        ("To review", GAP_STATUS_COLOR["unknown"]),
    ]:
        tr.conditional_formatting.add(
            f"J2:J400",
            FormulaRule(formula=[f'EXACT($J2,"{label}")'],
                        font=Font(name=FONT_SEMI, size=9, color=color)),
        )

    # ── 3. Summary ───────────────────────────────────────────────────────────
    sm = wb.create_sheet("Summary")
    sheet_setup(sm)
    set_widths(sm, [26, 13, 13, 15, 14, 10, 3, 22, 12])
    r = title_block(
        sm,
        "Where the gaps sit",
        "Findings per report chapter and gap type, and the live status of the list. "
        "All counts are COUNTIFS formulas over the “Gap tracker” sheet (rows 2–400), so they follow your edits.",
        width=9,
    )

    type_labels = [type_meta[t]["label"] for t in type_order]
    header_row(sm, r, ["Sector / report chapter"] + type_labels + ["Total"], height=32)
    first_data = r + 1
    for i, sector in enumerate(sectors):
        row = first_data + i
        vals = [sector]
        for t in type_order:
            vals.append(f'=COUNTIFS(\'{TRACKER}\'!$B$2:$B$400,$A{row},\'{TRACKER}\'!$C$2:$C$400,{get_column_letter(2 + type_order.index(t))}${r})')
        vals.append(f"=SUM(B{row}:E{row})")
        body_row(sm, row, vals, band=(i % 2 == 0), height=17)
        sm.cell(row=row, column=1).font = Font(name=FONT_SEMI, size=9, color=INK)
        for col in range(2, 7):
            sm.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
    total_row = first_data + len(sectors)
    body_row(sm, total_row, ["All chapters"] + [
        f"=SUM({get_column_letter(c)}{first_data}:{get_column_letter(c)}{total_row - 1})"
        for c in range(2, 7)
    ], band=True, height=18)
    for col in range(1, 7):
        c = sm.cell(row=total_row, column=col)
        c.font = Font(name=FONT_SEMI, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor=TEAL)
        if col > 1:
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Stacked bar of chapters × gap type.
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    data_ref = Reference(sm, min_col=2, max_col=5, min_row=r, max_row=total_row - 1)
    cats = Reference(sm, min_col=1, min_row=first_data, max_row=total_row - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    style_chart(chart, "Findings per report chapter, by gap type", height=10.5, width=18)
    text_categories(chart, sm, 1, first_data, total_row - 1)
    chart.y_axis.title = None
    chart.x_axis.title = "Number of findings"
    for series, t in zip(chart.series, type_order):
        series.graphicalProperties.solidFill = GAP_TYPE_COLOR[t]
        series.graphicalProperties.line.noFill = True
    chart.legend.position = "b"
    sm.add_chart(chart, f"H{r}")

    # Live-status block, below the matrix.
    r2 = total_row + 3
    sm.cell(row=r2, column=1, value="Live status of the list").font = H2_FONT
    r2 += 1
    header_row(sm, r2, ["Live status", "Findings", "", "", "", "", "", "", ""], height=20)
    status_first = r2 + 1
    status_labels = [status_meta[s]["label"] for s in status_meta]
    for i, (key, m) in enumerate(status_meta.items()):
        row = status_first + i
        body_row(sm, row, [m["label"],
                           f'=COUNTIF(\'{TRACKER}\'!$J$2:$J$400,$A{row})'], band=(i % 2 == 0), height=17)
        sm.cell(row=row, column=1).font = Font(name=FONT_SEMI, size=9, color=GAP_STATUS_COLOR[key])
        sm.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    status_last = status_first + len(status_meta) - 1

    pie = BarChart()
    pie.type = "col"
    pie.add_data(Reference(sm, min_col=2, min_row=r2, max_row=status_last), titles_from_data=True)
    pie.set_categories(Reference(sm, min_col=1, min_row=status_first, max_row=status_last))
    style_chart(pie, "Findings by live status", height=7.5, width=13)
    text_categories(pie, sm, 1, status_first, status_last)
    pie.legend = None
    pie.y_axis.title = "Number of findings"
    ser = pie.series[0]
    from openpyxl.chart.marker import DataPoint
    ser.data_points = [
        DataPoint(idx=i) for i in range(len(status_meta))
    ]
    for i, key in enumerate(status_meta):
        ser.data_points[i].graphicalProperties.solidFill = GAP_STATUS_COLOR[key]
        ser.data_points[i].graphicalProperties.line.noFill = True
    sm.add_chart(pie, f"H{max(r2 - 1, r + rows_for(10.5))}")

    note_line(sm, status_last + 2,
              "Note: all rows are seeded to the report's January-2024 baseline (“Still open”). A count against any other "
              "status means someone has reassessed that row — the reason should be in its “Status note”.",
              width=6)
    note_line(sm, status_last + 3,
              f"Source: {meta['author']} (2024) “{meta['title']}”, {meta['published']}; Method Hub Policy Gap Tracker.",
              width=6)

    # ── 4. Legend ────────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    sheet_setup(lg)
    set_widths(lg, [24, 74, 14])
    r = title_block(lg, "Legend", "The report's own classification (Ch. 2, methodology) and the live-status "
                                 "vocabulary used in the tracker.", width=3)
    header_row(lg, r, ["Gap type", "What it means (report methodology)", "Colour"], height=20)
    r += 1
    for i, (key, m) in enumerate(type_meta.items()):
        body_row(lg, r, [m["label"], m["description"], ""], band=(i % 2 == 0), height=20)
        lg.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=GAP_TYPE_COLOR[key])
        lg.cell(row=r, column=3).fill = PatternFill("solid", fgColor=GAP_TYPE_COLOR[key])
        r += 1
    r += 2
    header_row(lg, r, ["Live status", "When to use it", "Colour"], height=20)
    r += 1
    status_help = {
        "open": "Nothing adopted since January 2024 changes the finding — the report's baseline for every row.",
        "partially-addressed": "Legislation or implementation since the report narrows the gap but does not close it. "
                               "Say which instrument, and what is still missing, in the status note.",
        "addressed": "The gap is closed by an adopted instrument that is in force. Name the instrument in the note.",
        "unknown": "Needs a look — the row was flagged but not yet reassessed.",
    }
    for i, (key, m) in enumerate(status_meta.items()):
        body_row(lg, r, [m["label"], status_help[key], ""], band=(i % 2 == 0), height=28)
        lg.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=GAP_STATUS_COLOR[key])
        lg.cell(row=r, column=3).fill = PatternFill("solid", fgColor=GAP_STATUS_COLOR[key])
        r += 1
    r += 2
    note_line(lg, r, "Colours follow the ESABCC template palette (teal #4D7E83 with the template's accent colours), "
                     "so this workbook sits alongside the Board's Word documents.", width=3)

    # No cached results are written for the COUNTIFS formulas, so ask the
    # spreadsheet app to calculate the whole book the moment it opens.
    wb.calculation.fullCalcOnLoad = True
    wb.properties.title = "ESABCC Policy Gap Tracker — 2024 progress report"
    wb.properties.subject = "Summer Prep · Policy Gap 2.0 Report — background document"
    wb.properties.creator = "ESABCC Method Hub"
    fix_title_overlays(wb)
    wb.save(out_path)
    print(f"wrote {out_path}: {len(gaps)} findings, {len(sectors)} chapters")


if __name__ == "__main__":
    build(json.load(open(sys.argv[1], encoding="utf8")), sys.argv[2])
