"""
The combined workbook's opening page: the indicator module at a glance.

Everything on it is already somewhere else in the book — this sheet is the
top-down read of it. Six key figures, the largest moves as a diverging bar,
chapter coverage, four headline trend lines, and a latest-vs-target panel for
the four indicators the report gives an explicit numeric target.

The charts cannot reference the tables they come from directly (those are
spread over ten sheets with a different shape each), so their source blocks are
laid out on a second, hidden sheet — "Dashboard data" — one block per figure.
Those blocks are marked for the Derivations wiring like every other value cell,
so editing an input on Derivations moves the dashboard too.

Called by build_indicator_combined.py after the rest of the workbook exists;
it is not a standalone document.
"""

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from esabcc_style import (
    ACCENT_ORANGE, FONT, FONT_LIGHT, FONT_SEMI, H2_FONT, MUTED, NOTE_FONT, SUBTITLE_FONT,
    TEAL, TEAL_LIGHT, TITLE_FONT, WHITE, header_row, note_line, set_widths, sheet_setup,
    style_chart, text_categories,
)

DATA_SHEET = "Dashboard data"

# The four series worth a headline trend: the two headline aggregates, the
# electricity-mix share that carries the supply story, and transport — the one
# sector whose emissions are still above 2005.
TRENDS = [
    {"title": "O1 — Total EU GHG emissions", "codes": ["O1"]},
    {"title": "E2 — Non-biomass RES share of electricity", "codes": ["E2 (RES)"]},
    {"title": "O2 — Primary & final energy consumption",
     "codes": ["O2 (PEC)", "O2 (FEC)"], "labels": ["PEC", "FEC"]},
    {"title": "T1 — Transport GHG emissions", "codes": ["T1"]},
]

# Indicators whose target is a single number the latest value can be put next
# to. The other targets in the database are shares or per-head intensities
# whose "distance" needs the unit spelled out to mean anything, so they are
# left to the chapter tabs rather than reduced to a two-bar chart here.
TARGET_CODES = ["O1", "I3", "T5b", "B6"]

# Chart geometry. A chart is anchored to a cell but floats above the grid at
# its own size, so two charts side by side collide unless the gap between
# their anchor columns is WIDER than the left one. A default column is about
# 1.69 cm at 100 % zoom, so an 8-column gap is ~13.5 cm — which is why each
# width below is set under its stride rather than to a round number.
MOVES_SIZE = (17, 10.5)      # two figures across the sheet, 16 columns apart
TREND_SIZE = (14, 7.5)       # two across, 10 columns apart (~17 cm)
TREND_STRIDE = 10
TREND_ROW_STRIDE = 15        # ~7.9 cm of rows, so 7.5 cm of chart clears it
TARGET_SIZE = (12, 7)        # four across, 8 columns apart (~13.5 cm)
TARGET_STRIDE = 8

KPI_LABEL_FONT = Font(name=FONT_SEMI, size=9, color=WHITE)
KPI_VALUE_FONT = Font(name=FONT_SEMI, size=18, color=WHITE, bold=True)
KPI_FILL = PatternFill("solid", fgColor=TEAL)
RIGHT = Alignment(horizontal="right", vertical="center")


MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


def _stamp(prepared):
    """"27 July 2026" -> "2026-07", which is what fits in a key-figure tile."""
    parts = prepared.split()
    if len(parts) == 3 and parts[1] in MONTHS:
        return f"{parts[2]}-{MONTHS.index(parts[1]) + 1:02d}"
    return prepared


def _by_code(inds):
    return {i.get("code"): i for i in inds if i.get("code")}


def _kpi(ws, row, col, label, value):
    """One key-figure tile: label over value, five columns wide."""
    for r, text, font in ((row, label, KPI_LABEL_FONT), (row + 1, value, KPI_VALUE_FONT)):
        c = ws.cell(row=r, column=col, value=text)
        c.font = font
        c.fill = KPI_FILL
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=(r == row))
        for k in range(col, col + 4):
            ws.cell(row=r, column=k).fill = KPI_FILL
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 3)


def build_dashboard(inds, reads, factcheck, backed, wb, prepared, links=None):
    """
    Write the Dashboard and its hidden source sheet into `wb`.

    `reads` is build_indicator_check.read_indicator's output per id, `backed`
    the number of indicators that have a Derivations block, `prepared` the
    compilation date string used across the workbook.
    """
    dd = wb.create_sheet(DATA_SHEET)
    dd.sheet_state = "hidden"
    ws = wb.create_sheet("Dashboard")
    sheet_setup(ws)
    set_widths(ws, [13])

    updated = [i for i in inds if reads[i["id"]]["post"]]
    confirmed = sum(1 for r in factcheck.get("rows", []) if r.get("verdict") == "CONFIRMED")

    # ── title ───────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Dashboard — the indicator module at a glance").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    ws.row_dimensions[1].height = 26.1
    c = ws.cell(row=2, column=1, value=(
        f"The same {len(inds)} progress indicators as the rest of this workbook, read top-down instead of "
        "table-first: what has moved since the report, by how much, and where the derivation behind a figure "
        "is a live formula rather than a pasted number."))
    c.font = SUBTITLE_FONT
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    ws.row_dimensions[2].height = 30
    c = ws.cell(row=3, column=1, value=f"Summer Prep · Policy Gap 2.0 Report — prepared {prepared}.")
    c.font = Font(name=FONT_LIGHT, size=8, color=MUTED)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=20)

    # ── key figures ─────────────────────────────────────────────────────────
    share = round(100 * len(updated) / len(inds)) if inds else 0
    kpis = [
        ("Progress indicators", len(inds)),
        ("With post-report data", len(updated)),
        ("Share of series moved on", f"{share}%"),
        ("Derivation-backed indicators", backed),
        ("Confirmed fact-check points", confirmed),
        ("Data as of", _stamp(prepared)),
    ]
    for n, (label, value) in enumerate(kpis):
        _kpi(ws, 5, 1 + 5 * n, label, value)
    ws.row_dimensions[5].height = 15.95
    ws.row_dimensions[6].height = 26.1

    # ── the source blocks, then the figures over them ───────────────────────
    _moves_block(dd, inds, reads, links)
    _coverage_block(dd, inds, reads)
    trend_cols = _trend_blocks(dd, inds, links)
    target_cols = _target_blocks(dd, inds, reads, links)

    ws.cell(row=8, column=1, value="Largest moves since the report").font = H2_FONT
    ws.cell(row=8, column=17, value="Chapter coverage").font = H2_FONT
    ws.row_dimensions[8].height = 17.25
    _moves_chart(ws, dd, inds, reads)
    _coverage_chart(ws, dd, inds, reads)

    ws.cell(row=53, column=1, value="Headline trends").font = H2_FONT
    ws.row_dimensions[53].height = 17.25
    for n, (spec, cols) in enumerate(trend_cols):
        anchor = f"{get_column_letter(1 + TREND_STRIDE * (n % 2))}{54 + TREND_ROW_STRIDE * (n // 2)}"
        _trend_chart(ws, dd, spec, cols, anchor)

    ws.cell(row=85, column=1, value="Distance to target").font = H2_FONT
    ws.row_dimensions[85].height = 17.25
    for n, spec in enumerate(target_cols):
        _target_chart(ws, dd, spec, f"{get_column_letter(1 + TARGET_STRIDE * n)}86")

    note_line(ws, 101, (
        "Notes: “moved on” means the indicator has gained at least one data point since the 2024 report. "
        "Figures fed from a Derivations block are live formulas, exactly as in the rest of this workbook — edit "
        "an input there and this page follows on the next recalculation. Nothing here is a Board position."),
        width=20)
    return ws


# ── source blocks ───────────────────────────────────────────────────────────
def _moves_block(dd, inds, reads, links):
    """Top 8 up and top 8 down, smallest first so the bar chart reads bottom-up."""
    movers = [i for i in inds if reads[i["id"]]["post"] and reads[i["id"]]["pct"] is not None]
    movers.sort(key=lambda i: reads[i["id"]]["pct"])
    picked = movers[:8] + movers[-8:] if len(movers) > 16 else movers
    picked.sort(key=lambda i: reads[i["id"]]["pct"])

    dd.cell(row=1, column=1, value="Largest moves — code · name")
    dd.cell(row=1, column=2, value="Change %")
    dd.cell(row=1, column=8, value="Latest value")
    for n, ind in enumerate(picked):
        rd = reads[ind["id"]]
        row = 2 + n
        dd.cell(row=row, column=1, value=f"{ind.get('code')} · {ind['name']}")
        dd.cell(row=row, column=2, value=rd["pct"] / 100)
        dd.cell(row=row, column=8, value=rd["latest"]["value"])
        if links:
            links.mark(dd, row, 8, ind["id"], rd["latest"]["year"])
    return picked


def _coverage_block(dd, inds, reads):
    from build_indicator_check import CATEGORY_LABEL, CATEGORY_ORDER

    counts = {}
    for i in inds:
        up, still = counts.get(i["category"], (0, 0))
        counts[i["category"]] = (up + 1, still) if reads[i["id"]]["post"] else (up, still + 1)
    cats = [c for c in CATEGORY_ORDER if c in counts] + \
           [c for c in counts if c not in CATEGORY_ORDER]

    dd.cell(row=1, column=4, value="Chapter")
    dd.cell(row=1, column=5, value="Moved on")
    dd.cell(row=1, column=6, value="Still at report")
    for n, cat in enumerate(cats):
        up, still = counts[cat]
        dd.cell(row=2 + n, column=4, value=CATEGORY_LABEL.get(cat, cat))
        dd.cell(row=2 + n, column=5, value=up)
        dd.cell(row=2 + n, column=6, value=still)
    return cats


def _trend_blocks(dd, inds, links):
    """
    One three-column block per headline trend: Year, then two series.

    A single indicator is split into "in the report" and "added since", with
    the join year repeated in both so the two lines meet. A two-indicator
    trend (O2's primary and final energy) uses one column each instead.
    Blocks are laid out from column J with a one-column gap between them.
    """
    by_code = _by_code(inds)
    out = []
    col = 10
    for spec in TRENDS:
        chosen = [by_code[c] for c in spec["codes"] if c in by_code]
        if not chosen:
            continue
        years = sorted({d["year"] for i in chosen for d in i["data"]})
        dd.cell(row=1, column=col, value="Year")
        if len(chosen) == 2:
            headers = spec.get("labels") or [i.get("code") for i in chosen]
            series = [{d["year"]: d for d in i["data"]} for i in chosen]
            owners = chosen
        else:
            ind = chosen[0]
            headers = ["In the 2024 report", "Added since the report"]
            pre = {d["year"]: d for d in ind["data"] if not d.get("afterReport")}
            post = {d["year"]: d for d in ind["data"] if d.get("afterReport")}
            if pre and post:            # repeat the last report point, so the lines join
                join = max(pre)
                post = {**post, join: pre[join]}
            series = [pre, post]
            owners = [ind, ind]
        for k, h in enumerate(headers):
            dd.cell(row=1, column=col + 1 + k, value=h)
        for n, year in enumerate(years):
            row = 2 + n
            dd.cell(row=row, column=col, value=year)
            for k, points in enumerate(series):
                d = points.get(year)
                if d is None:
                    continue
                dd.cell(row=row, column=col + 1 + k, value=d["value"])
                if links:
                    links.mark(dd, row, col + 1 + k, owners[k]["id"], year)
        out.append((dict(spec, indicator=chosen[0]),
                    {"col": col, "first": 2, "last": 1 + len(years), "series": len(series)}))
        col += 4
    return out


def _target_blocks(dd, inds, reads, links):
    """Latest value against the target, one two-row block per indicator."""
    by_code = _by_code(inds)
    out = []
    col = 26
    for code in TARGET_CODES:
        ind = by_code.get(code)
        if not ind or ind.get("targetValue") is None:
            continue
        rd = reads[ind["id"]]
        if not rd["latest"]:
            continue
        dd.cell(row=1, column=col, value=code)
        dd.cell(row=2, column=col, value="Latest")
        dd.cell(row=3, column=col, value="Target")
        dd.cell(row=2, column=col + 1, value=rd["latest"]["value"])
        dd.cell(row=3, column=col + 1, value=ind["targetValue"])
        if links:
            links.mark(dd, 2, col + 1, ind["id"], rd["latest"]["year"])
        out.append({"ind": ind, "col": col})
        col += 2
    return out


# ── figures ─────────────────────────────────────────────────────────────────
def _moves_chart(ws, dd, inds, reads):
    rows = sum(1 for r in dd.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True) if r[0])
    if not rows:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.add_data(Reference(dd, min_col=2, min_row=1, max_row=1 + rows), titles_from_data=True)
    chart.set_categories(Reference(dd, min_col=1, min_row=2, max_row=1 + rows))
    style_chart(chart, "Change vs the report's last figure — top 8 up, top 8 down",
                width=MOVES_SIZE[0], height=MOVES_SIZE[1])
    text_categories(chart, dd, 1, 2, 1 + rows)
    chart.legend = None
    chart.x_axis.numFmt = "0%"
    ser = chart.series[0]
    ser.data_points = [DataPoint(idx=i) for i in range(rows)]
    for i in range(rows):
        pct = dd.cell(row=2 + i, column=2).value or 0
        ser.data_points[i].graphicalProperties.solidFill = TEAL if pct < 0 else ACCENT_ORANGE
        ser.data_points[i].graphicalProperties.line.noFill = True
    ws.add_chart(chart, "A9")


def _coverage_chart(ws, dd, inds, reads):
    rows = sum(1 for r in dd.iter_rows(min_col=4, max_col=4, min_row=2, values_only=True) if r[0])
    if not rows:
        return
    cov = BarChart()
    cov.type = "col"
    cov.grouping = "stacked"
    cov.overlap = 100
    cov.add_data(Reference(dd, min_col=5, max_col=6, min_row=1, max_row=1 + rows),
                 titles_from_data=True)
    cov.set_categories(Reference(dd, min_col=4, min_row=2, max_row=1 + rows))
    style_chart(cov, "Indicators moved on vs still at the report figure, by chapter",
                width=MOVES_SIZE[0], height=MOVES_SIZE[1])
    text_categories(cov, dd, 4, 2, 1 + rows)
    cov.y_axis.title = "Indicators"
    for series, color in zip(cov.series, [TEAL, TEAL_LIGHT]):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.noFill = True
    cov.legend.position = "b"
    ws.add_chart(cov, "Q9")


def _trend_chart(ws, dd, spec, cols, anchor):
    col, first, last = cols["col"], cols["first"], cols["last"]
    ch = LineChart()
    ch.add_data(Reference(dd, min_col=col + 1, max_col=col + cols["series"],
                          min_row=1, max_row=last), titles_from_data=True)
    ch.set_categories(Reference(dd, min_col=col, min_row=first, max_row=last))
    style_chart(ch, f"{spec['title']} ({spec['indicator']['unit']})",
                width=TREND_SIZE[0], height=TREND_SIZE[1])
    ch.y_axis.title = spec["indicator"]["unit"]
    for k, (series, color) in enumerate(zip(ch.series, [TEAL, ACCENT_ORANGE])):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 22000
        series.smooth = False
        series.marker = Marker(symbol="none")
        if k == 1 and "labels" not in spec:
            series.graphicalProperties.line.dashStyle = "dash"
            series.marker = Marker(symbol="circle", size=5)
    ch.legend.position = "b"
    ws.add_chart(ch, anchor)


def _target_chart(ws, dd, spec, anchor):
    ind, col = spec["ind"], spec["col"]
    ch = BarChart()
    ch.type = "bar"
    ch.add_data(Reference(dd, min_col=col + 1, min_row=2, max_row=3))
    ch.set_categories(Reference(dd, min_col=col, min_row=2, max_row=3))
    style_chart(ch, f"{ind['code']} · latest vs {ind['targetYear']} target ({ind['unit']})",
                width=TARGET_SIZE[0], height=TARGET_SIZE[1])
    text_categories(ch, dd, col, 2, 3)
    ch.legend = None
    ser = ch.series[0]
    ser.data_points = [DataPoint(idx=0), DataPoint(idx=1)]
    for dp, color in zip(ser.data_points, [TEAL, ACCENT_ORANGE]):
        dp.graphicalProperties.solidFill = color
        dp.graphicalProperties.line.noFill = True
    ws.add_chart(ch, anchor)
