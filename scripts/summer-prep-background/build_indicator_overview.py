"""
Background documents 5 and 6 — the two indicator overviews.

  A. "New data overview" — one big figure carrying every data point added since
     the 2024 report. Indicators are in wildly different units, so the only
     honest way to put 97 series on one chart is to index each to its own report
     baseline (= 100): the figure then reads as "how far, and which way, has each
     indicator moved since the report", which is the question the module asks.
     The underlying numbers sit next to the figure, so nothing is hidden in it.

  B. "Old indicators, new data" — every indicator the 2024 report carried, its
     report baseline and its latest value, with the DERIVATION written out as a
     live Excel formula. Open a cell and Excel shows =IF(A21>=2023, ROUND(...))
     over the raw inputs, exactly as the Method Hub's calc grid computes it, so
     the arithmetic behind each number is visible and re-computable in the sheet.

Where the report's own derivation can be carried forward — its own input
columns refreshed from the same publisher, its own formula untouched — that is
what the Derivations sheet shows, and the sheet says so per indicator. Where it
cannot (the inputs are a pocketbook table or an association's fleet count with
no feed), the sheet says that too rather than passing off a substitute recipe
as the report's method.

Usage:
  python3 build_indicator_overview.py data.json calc_excel.json factcheck.json reportway.json outdir
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from esabcc_style import (
    ACCENT_ORANGE, ACCENT_PURPLE, ACCENT_RED, BODY_FILL, BODY_FONT, FONT, FONT_LIGHT,
    FONT_SEMI, H2_FONT, INK, LABEL_FONT, MUTED, SMALL_FONT, TEAL, TEAL_PALE, WHITE,
    body_row, header_row, note_line, set_widths, sheet_setup, style_chart,
    text_categories, title_block,
)
from dataset_links import detect_link, link_cell

PREPARED = "27 July 2026"

CATEGORY_LABEL = {
    "emissions": "Emissions", "energy-supply": "Energy supply",
    "energy-demand": "Energy demand", "transport": "Transport",
    "buildings": "Buildings", "industry": "Industry", "agriculture": "Agriculture",
    "lulucf": "LULUCF", "finance": "Finance & innovation", "fairness": "Fairness",
    "adaptation": "Adaptation",
}
CATEGORY_ORDER = ["emissions", "energy-supply", "energy-demand", "industry", "transport",
                  "buildings", "agriculture", "lulucf", "finance", "fairness", "adaptation"]
CATEGORY_COLOR = {
    "emissions": TEAL, "energy-supply": "3A6165", "energy-demand": "648ECA",
    "transport": ACCENT_PURPLE, "buildings": "00928F", "industry": ACCENT_RED,
    "agriculture": "6E8B1F", "lulucf": "007B6C", "finance": "A530B8",
    "fairness": ACCENT_ORANGE, "adaptation": "478EA5",
}
VERDICT_SHORT = {
    "CONFIRMED": "Reproduces from source",
    "REVISION": "2–5% off source",
    "WRONG": "Does not reproduce",
    "NO SOURCE YEAR": "No source figure yet",
    "NOT CHECKABLE": "Published figure, not re-derived",
}


def read(ind):
    data = sorted(ind["data"], key=lambda d: d["year"])
    pre = [d for d in data if not d.get("afterReport")]
    post = [d for d in data if d.get("afterReport")]
    base = pre[-1] if pre else None
    latest = data[-1] if data else None
    pct = None
    if base and latest and base["value"] and post:
        pct = (latest["value"] - base["value"]) / abs(base["value"]) * 100
    return {"data": data, "pre": pre, "post": post, "baseline": base,
            "latest": latest, "pct": pct}


# ── A. the one big figure ───────────────────────────────────────────────────

def build_new_data_overview(inds, reads, fc, out_path=None, wb=None, title="The one big figure"):
    """
    Render the "everything that has moved" figure.

    Standalone (default): `wb` is None, a fresh Workbook is created and
    saved to `out_path`. Embedded: pass an existing `wb` — the sheet is
    appended under `title` (the combined workbook uses "New data") and
    nothing is saved here.
    """
    standalone = wb is None
    if standalone:
        wb = Workbook()
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    sheet_setup(ws)
    set_widths(ws, [10, 44, 17, 13, 11, 13, 11, 13, 12, 13, 26])

    updated = [i for i in inds if reads[i["id"]]["post"] and reads[i["id"]]["pct"] is not None]
    updated.sort(key=lambda i: reads[i["id"]]["pct"])

    r = title_block(
        ws,
        "Everything that has moved since the 2024 report, in one figure",
        f"{len(updated)} of the {len(inds)} progress indicators have gained data since the report. "
        "They are measured in Mt, TWh, %, vehicles and head of cattle, so they are shown as the change "
        "against each indicator's own report baseline — the figure answers “how far, and which way”, "
        "not “how big”. The numbers behind every bar are in the table below it.",
        f"Summer Prep · Policy Gap 2.0 Report — prepared {PREPARED}. "
        "No better/worse reading is applied: whether a move is progress depends on the indicator.",
        width=11,
    )

    header_row(ws, r, [
        "Code", "Indicator", "Chapter", "Unit", "Baseline yr", "Baseline value",
        "Latest yr", "Latest value", "Change", "Change %", "Source check (27 Jul 2026)",
    ], height=32)
    head = r
    first = r + 1
    for n, ind in enumerate(updated):
        rd = reads[ind["id"]]
        b, l = rd["baseline"], rd["latest"]
        row = first + n
        v = fc.get(ind["id"], {}).get("verdict")
        body_row(ws, row, [
            ind.get("code"), ind["name"], CATEGORY_LABEL.get(ind["category"], ind["category"]),
            ind["unit"], b["year"], b["value"], l["year"], l["value"],
            l["value"] - b["value"], rd["pct"] / 100,
            VERDICT_SHORT.get(v, "—"),
        ], band=(n % 2 == 0), height=15)
        ws.cell(row=row, column=1).font = Font(name=FONT_SEMI, size=9,
                                               color=CATEGORY_COLOR.get(ind["category"], TEAL))
        for c in (5, 7):
            ws.cell(row=row, column=c).number_format = "0"
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")
        for c in (6, 8, 9):
            ws.cell(row=row, column=c).number_format = "#,##0.00"
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=10).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(row=row, column=10).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=11).font = SMALL_FONT
    last = first + len(updated) - 1
    ws.auto_filter.ref = f"A{head}:K{last}"

    # The figure itself: one bar per indicator, tall enough to name every one.
    chart = BarChart()
    chart.type = "bar"
    chart.add_data(Reference(ws, min_col=10, min_row=head, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=2, min_row=first, max_row=last))
    style_chart(chart, "Change since the report baseline — every indicator with new data",
                height=1.05 * len(updated) + 4, width=26)
    text_categories(chart, ws, 2, first, last)
    chart.legend = None
    chart.x_axis.title = "Change vs the report's last figure (%)"
    chart.x_axis.numFmt = "0%"
    ser = chart.series[0]
    ser.data_points = [DataPoint(idx=i) for i in range(len(updated))]
    for i, ind in enumerate(updated):
        ser.data_points[i].graphicalProperties.solidFill = CATEGORY_COLOR.get(ind["category"], TEAL)
        ser.data_points[i].graphicalProperties.line.noFill = True
    ws.add_chart(chart, f"M{head}")

    note_line(ws, last + 2,
              "Notes: bars are coloured by report chapter. A long bar is a large move relative to that "
              "indicator's own baseline, not a large quantity — indicators with small baselines (capacity "
              "additions, ZEV shares) move by large percentages by nature.", width=11)
    note_line(ws, last + 3,
              "Source: ESABCC (2024) “Towards EU climate neutrality”, progress indicators, extended with the "
              "post-publication points in the Policy Gap 2.0 indicator database. Every post-report value was "
              "re-checked against its primary source on 27 July 2026 (see the “Source check” column).",
              width=11)

    # A second view: the new points per chapter, so coverage is visible too.
    by_cat = defaultdict(lambda: [0, 0])
    for i in inds:
        by_cat[i["category"]][0 if reads[i["id"]]["post"] else 1] += 1
    cats = [c for c in CATEGORY_ORDER if c in by_cat]
    cr = last + 5
    ws.cell(row=cr, column=1, value="Coverage — how much of each chapter has new data").font = H2_FONT
    cr += 1
    header_row(ws, cr, ["", "Chapter", "With new data", "Still at the report figure",
                        "", "", "", "", "", "", ""], height=22)
    c_first = cr + 1
    for n, cat in enumerate(cats):
        up, still = by_cat[cat]
        body_row(ws, c_first + n, ["", CATEGORY_LABEL.get(cat, cat), up, still],
                 band=(n % 2 == 0), height=15)
        for c in (3, 4):
            ws.cell(row=c_first + n, column=c).alignment = Alignment(horizontal="center", vertical="center")
    c_last = c_first + len(cats) - 1
    cov = BarChart()
    cov.type = "col"
    cov.grouping = "stacked"
    cov.overlap = 100
    cov.add_data(Reference(ws, min_col=3, max_col=4, min_row=cr, max_row=c_last), titles_from_data=True)
    cov.set_categories(Reference(ws, min_col=2, min_row=c_first, max_row=c_last))
    style_chart(cov, "Indicators with new data since the report, by chapter", height=9, width=20)
    text_categories(cov, ws, 2, c_first, c_last)
    for series, color in zip(cov.series, [TEAL, "D8E7E8"]):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.noFill = True
    cov.legend.position = "b"
    cov.y_axis.title = "Indicators"
    ws.add_chart(cov, f"F{cr}")

    if standalone:
        wb.calculation.fullCalcOnLoad = True
        wb.properties.title = "ESABCC indicators — everything that has moved since the 2024 report"
        wb.properties.creator = "ESABCC Method Hub"
        wb.save(out_path)
        print(f"wrote {out_path}: one figure over {len(updated)} updated indicators")
    return wb


# ── B. old indicators + new data + the derivation as a live formula ─────────

DERIVATION_KIND = {
    "report": "Report's own derivation, continued",
    "switched": "Different recipe for the later years",
    "reconstructed": "Report's formula, input back-computed",
    "published": "Published figure, no derivation",
    "none": "No calc grid",
}


def derivation_kind(iid, calc, reportway):
    """How the later years of this indicator are obtained."""
    if iid in reportway:
        return "report"
    layout = calc.get(iid)
    if not layout:
        return "none"
    headers = [c["header"] for c in layout["columns"]]
    if any(h.startswith("Post-report figure as published") for h in headers):
        return "published"
    expr = layout["columns"][0].get("expr") or ""
    return "switched" if expr.startswith("IF(") else "reconstructed"


def build_old_vs_new(inds, reads, calc, fc, reportway, out_path=None, wb=None):
    """
    Render the Old-vs-New-with-derivations workbook.

    Standalone (default): `wb` is None, a fresh Workbook is created, the
    sheet's own "Read me" is written, and the result is saved to `out_path`.

    Embedded (combined workbook): pass an existing `wb` — its sheets are
    appended to (no "Read me" of its own), and nothing is saved here.
    `out_path` is ignored.
    """
    standalone = wb is None
    if standalone:
        wb = Workbook()

        ws = wb.active
        ws.title = "Read me"
        sheet_setup(ws)
        set_widths(ws, [24, 22, 22, 22, 22, 20, 16, 16])
        r = title_block(
            ws,
            "The report's indicators, and where they stand now",
            "Every progress indicator the 2024 report carried, with the figure the report itself published, "
            "the latest figure in the Policy Gap 2.0 workspace database, and the arithmetic in between — "
            "written into the sheet as a live Excel formula.",
            f"Summer Prep · Policy Gap 2.0 Report — prepared {PREPARED}.",
        )
        ws.cell(row=r, column=1, value="What is in this workbook").font = H2_FONT
        r += 1
        for i, (name, desc) in enumerate([
            ("Old vs new", "One row per indicator: what the report published, what the database holds now, the "
                           "change, and how that later figure was obtained. Sorted by report chapter."),
            ("Derivations", "The calc grid behind every indicator, one block each: Year, Value and the input "
                            "columns. The Value cells are LIVE EXCEL FORMULAS over the inputs in the same block — "
                            "click one and Excel shows the arithmetic. Change an input and the value recomputes, "
                            "exactly as the Method Hub's calc editor does it."),
            ("Sources", "The exact source query behind every input column — dataset, filters and URL."),
        ]):
            ws.cell(row=r, column=1, value=name).font = Font(name=FONT_SEMI, size=9, color=TEAL)
            ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
            c = ws.cell(row=r, column=2, value=desc)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = BODY_FILL if i % 2 == 0 else PatternFill("solid", fgColor=TEAL_PALE)
            ws.row_dimensions[r].height = 46
            r += 1
        r += 1
        for line in [
            "• The formulas are the Method Hub's own: they are produced by the same translator the Hub uses to "
            "export a calc grid to Excel, so what you see here is what the Hub computes, not a re-implementation.",
            "• A formula of the form =IF(A21>=2023, ROUND(<live-source recipe>), <report recipe>) means the report "
            "years keep the 2024 workbook's derivation and the later years are computed from the live source. That "
            "year switch is deliberate — the two are different derivations, and both are shown.",
            "• Where an indicator is a published figure with no derivation, the input column simply carries that "
            "figure and names the publication. Nothing is left blank.",
            "• Every post-report value was re-checked against its primary source on 27 July 2026; the verdict is in "
            "the “Source check” column of the Old vs new sheet.",
        ]:
            r = note_line(ws, r, line, width=8, font=BODY_FONT)
            ws.row_dimensions[r - 1].height = 30

    # ── Old vs new ─────────────────────────────────────────────────────────
    ov = wb.create_sheet("Old vs new")
    sheet_setup(ov, freeze="C6")
    set_widths(ov, [11, 44, 16, 14, 12, 14, 11, 14, 12, 12, 24, 30, 34])
    r = title_block(
        ov, "What the report published, and what the database holds now",
        "“Report figure” is the last value the 2024 report itself carried. “Latest” is the most recent value in "
        "the Policy Gap 2.0 workspace indicator database today.", width=12)
    header_row(ov, r, [
        "Code", "Indicator", "Chapter", "Unit", "Report year", "Report figure",
        "Latest year", "Latest value", "Change", "Change %", "How the later figure was obtained",
        "Source check (27 Jul 2026)", "Primary source",
    ], height=32)
    head = r
    r += 1
    order = {c: n for n, c in enumerate(CATEGORY_ORDER)}
    for n, ind in enumerate(sorted(inds, key=lambda i: (order.get(i["category"], 99), i.get("code") or ""))):
        rd = reads[ind["id"]]
        b, l = rd["baseline"], rd["latest"]
        has_new = bool(rd["post"])
        kind = derivation_kind(ind["id"], calc, reportway)
        how = ("No later data — report figure stands" if not has_new
               else DERIVATION_KIND[kind])
        v = fc.get(ind["id"], {}).get("verdict")
        body_row(ov, r + n, [
            ind.get("code"), ind["name"], CATEGORY_LABEL.get(ind["category"], ind["category"]),
            ind["unit"],
            b["year"] if b else None, b["value"] if b else None,
            l["year"] if l else None, l["value"] if l else None,
            (l["value"] - b["value"]) if (b and l and has_new) else None,
            (rd["pct"] / 100) if rd["pct"] is not None else None,
            how, VERDICT_SHORT.get(v, "—") if has_new else "—",
            ind.get("source") or "",
        ], band=(n % 2 == 0), height=15)
        ov.cell(row=r + n, column=1).font = Font(name=FONT_SEMI, size=9,
                                                 color=CATEGORY_COLOR.get(ind["category"], TEAL))
        for c in (5, 7):
            ov.cell(row=r + n, column=c).number_format = "0"
            ov.cell(row=r + n, column=c).alignment = Alignment(horizontal="center", vertical="center")
        for c in (6, 8, 9):
            ov.cell(row=r + n, column=c).number_format = "#,##0.00"
            ov.cell(row=r + n, column=c).alignment = Alignment(horizontal="right", vertical="center")
        ov.cell(row=r + n, column=10).number_format = "+0.0%;-0.0%;0.0%"
        ov.cell(row=r + n, column=10).alignment = Alignment(horizontal="right", vertical="center")
        ov.cell(row=r + n, column=11).font = SMALL_FONT
        ov.cell(row=r + n, column=12).font = SMALL_FONT
        src_cell = ov.cell(row=r + n, column=13)
        src_cell.font = SMALL_FONT
        link_cell(src_cell, detect_link(ind.get("source")))
    ov.auto_filter.ref = f"A{head}:M{head + len(inds)}"

    # ── Derivations: the calc grid, with live formulas ─────────────────────
    dv = wb.create_sheet("Derivations")
    sheet_setup(dv)
    set_widths(dv, [10, 16, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22])
    r = title_block(
        dv, "The arithmetic behind every figure",
        "One block per indicator: Year, Value, then the inputs the value is built from. The Value cells are live "
        "Excel formulas over the cells to their right — click one to see it, edit an input and it recomputes. "
        "These are the Method Hub's own formulas, produced by the same translator the Hub uses to export a calc "
        "grid, so the sheet and the Hub agree by construction.",
        width=12)
    blocks = 0
    for ind in sorted(inds, key=lambda i: (order.get(i["category"], 99), i.get("code") or "")):
        kind = derivation_kind(ind["id"], calc, reportway)
        layout = reportway.get(ind["id"]) or calc.get(ind["id"])
        if not layout:
            continue
        blocks += 1
        title = f"{ind.get('code')} · {ind['name']}  ({ind['unit']})"
        c = dv.cell(row=r, column=1, value=title)
        c.font = Font(name=FONT_SEMI, size=11, color=CATEGORY_COLOR.get(ind["category"], TEAL))
        dv.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        label = {
            "report": "SAME DERIVATION AS THE REPORT — the report's own formula, with its own input "
                      "columns refreshed from the same publisher for the later years. Nothing about "
                      "the calculation changes across the join.",
            "switched": "NOT the report's derivation for the later years — the formula switches by "
                        "year to a different recipe, because the report's own inputs have no feed. "
                        "The two halves of this series are not computed the same way.",
            "reconstructed": "The report's formula throughout; for the later years its single input "
                             "is back-computed from the published figure through the sheet's own "
                             "constants, so the arithmetic is unchanged.",
            "published": "No derivation for the later years — the figure is the publication's own.",
            "none": "",
        }[kind]
        if label:
            c = dv.cell(row=r, column=1, value=label)
            c.font = Font(name=FONT_SEMI, size=8.5,
                          color=TEAL if kind in ("report", "reconstructed") else ACCENT_ORANGE)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            dv.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            dv.row_dimensions[r].height = 26
            r += 1
        rw = reportway.get(ind["id"])
        if rw and rw.get("refreshed"):
            names = ", ".join(f"{c['header']}" for c in rw["refreshed"])
            c = dv.cell(row=r, column=1, value=f"Refreshed for {', '.join(str(y) for y in rw['addedYears'])}: {names}")
            c.font = Font(name=FONT_LIGHT, size=8, color=MUTED)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            dv.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            dv.row_dimensions[r].height = 22
            r += 1
        expr = layout["columns"][0].get("expr")
        if expr:
            c = dv.cell(row=r, column=1, value=f"Derivation:  Value = {expr}")
            c.font = Font(name=FONT_LIGHT, size=8, color=MUTED)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            dv.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            dv.row_dimensions[r].height = 24
            r += 1

        headers = ["Year"] + [col["header"] for col in layout["columns"]]
        header_row(dv, r, headers, height=42)
        for i in range(1, len(headers) + 1):
            dv.cell(row=r, column=i).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        header_excel_row = r
        r += 1
        first_data = r
        for rowdata in layout["rows"]:
            vals = [rowdata["year"]]
            for ci, col in enumerate(layout["columns"]):
                # a derived column becomes a live formula; a raw input keeps its number
                f = rowdata["formulas"][ci] if ci < len(rowdata["formulas"]) else None
                if f:
                    vals.append(None)   # written below, offset for this block
                else:
                    cell = rowdata["cells"][ci] if ci < len(rowdata["cells"]) else None
                    if isinstance(cell, dict):
                        cell = cell.get("v")
                    vals.append(cell)
            body_row(dv, r, vals, band=((r - first_data) % 2 == 0), height=14)
            dv.cell(row=r, column=1).number_format = "0"
            dv.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
            for ci in range(len(layout["columns"])):
                f = rowdata["formulas"][ci] if ci < len(rowdata["formulas"]) else None
                cell = dv.cell(row=r, column=2 + ci)
                if f:
                    # the translator addressed row (index+2); shift onto this block's rows
                    shift = first_data - 2
                    cell.value = "=" + _shift_rows(f, shift)
                    cell.font = Font(name=FONT, size=9, color=INK, bold=(ci == 0))
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            r += 1
        r += 2
    note_line(dv, r, "Notes: a formula of the form =IF(A…>=2023, ROUND(<live-source recipe>), <report recipe>) "
                     "keeps the report years on the 2024 workbook's derivation and computes the later years from "
                     "the live source. Both derivations are shown because they are genuinely different.", width=12)

    # ── Sources ────────────────────────────────────────────────────────────
    sr = wb.create_sheet("Sources")
    sheet_setup(sr, freeze="A2")
    header_row(sr, 1, ["Code", "Indicator", "Input column", "Where that input comes from",
                       "Exact dataset link (DOI where available)"], height=22)
    set_widths(sr, [11, 40, 40, 100, 26])
    row = 2
    for ind in sorted(inds, key=lambda i: (order.get(i["category"], 99), i.get("code") or "")):
        layout = calc.get(ind["id"])
        if not layout:
            continue
        for ci, col in enumerate(layout["columns"]):
            if not col.get("source"):
                continue
            body_row(sr, row, [ind.get("code") if ci == 0 else "", ind["name"] if ci == 0 else "",
                               col["header"], col["source"]], band=(row % 2 == 0), height=26)
            sr.cell(row=row, column=1).font = Font(name=FONT_SEMI, size=9, color=TEAL)
            sr.cell(row=row, column=4).font = SMALL_FONT
            link_cell(sr.cell(row=row, column=5), detect_link(col["source"]))
            row += 1
    sr.auto_filter.ref = f"A1:E{row - 1}"

    if standalone:
        wb.calculation.fullCalcOnLoad = True
        wb.properties.title = "ESABCC report indicators — old figures, new data, and the derivations"
        wb.properties.creator = "ESABCC Method Hub"
        wb.save(out_path)
        print(f"wrote {out_path}: {len(inds)} indicators, {blocks} derivation blocks")
    return wb


def _shift_rows(formula, shift):
    """Move every A1 row reference in a formula down by `shift` rows."""
    import re
    if not shift:
        return formula
    return re.sub(r"\b([A-Z]{1,2})(\d+)\b",
                  lambda m: f"{m.group(1)}{int(m.group(2)) + shift}", formula)


def main(data_path, calc_path, fc_path, reportway_path, outdir):
    data = json.load(open(data_path, encoding="utf8"))
    calc = json.load(open(calc_path, encoding="utf8"))
    fcraw = json.load(open(fc_path, encoding="utf8"))
    order = ["CONFIRMED", "NOT CHECKABLE", "NO SOURCE YEAR", "REVISION", "WRONG"]
    fc = {}
    for row in fcraw.get("rows", []):
        cur = fc.setdefault(row["id"], {"verdict": "CONFIRMED"})
        if order.index(row["verdict"]) > order.index(cur["verdict"]):
            cur["verdict"] = row["verdict"]

    reportway = json.load(open(reportway_path, encoding="utf8"))["filled"]
    inds = data["indicators"]
    reads = {i["id"]: read(i) for i in inds}
    out = Path(outdir)
    build_new_data_overview(inds, reads, fc, out / "ESABCC_Indicator-New-Data-Overview_2026-07.xlsx")
    build_old_vs_new(inds, reads, calc, fc, reportway,
                     out / "ESABCC_Indicators-Old-vs-New-with-derivations_2026-07.xlsx")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5])
