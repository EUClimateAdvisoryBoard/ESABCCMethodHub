"""
Background document 2 — Indicator Check (all sectors).

"What has moved, data-wise, since the last report?" Renders the Summer Prep
Indicator Check (Note 1) as a standalone workbook: every ESABCC progress
indicator with its report baseline, the points added since publication, the
change between them, a figure per indicator, and — behind all of it — the calc
inputs each post-report number was built from, so no figure on a chart is
without a visible source.

Usage:  python3 build_indicator_check.py data.json calc.json factcheck.json out.xlsx
"""

import json
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from esabcc_style import (
    ACCENT_BLUE, ACCENT_ORANGE, ACCENT_PURPLE, ACCENT_RED, BODY_FILL, BODY_FONT, FONT,
    FONT_LIGHT, FONT_SEMI, H2_FONT, INK, LABEL_FONT, MUTED, RULE, SMALL_FONT, TEAL,
    TEAL_LIGHT, TEAL_PALE, WHITE, body_row, header_row, note_line, set_widths,
    sheet_setup, style_chart, text_categories, title_block,
)
from dataset_links import detect_link, link_cell

PREPARED = "27 July 2026"
REPORT_YEAR = "January 2024"

CATEGORY_LABEL = {
    "emissions": "Emissions",
    "energy-supply": "Energy supply",
    "energy-demand": "Energy demand",
    "transport": "Transport",
    "buildings": "Buildings",
    "industry": "Industry",
    "agriculture": "Agriculture",
    "lulucf": "LULUCF",
    "finance": "Finance & innovation",
    "fairness": "Fairness",
    "adaptation": "Adaptation",
}
# Chapter order of the report, so the tabs read like the report itself.
CATEGORY_ORDER = [
    "emissions", "energy-supply", "energy-demand", "industry", "transport",
    "buildings", "agriculture", "lulucf", "finance", "fairness", "adaptation",
]
CATEGORY_COLOR = {
    "emissions": TEAL,
    "energy-supply": "3A6165",
    "energy-demand": ACCENT_BLUE,
    "transport": ACCENT_PURPLE,
    "buildings": "00928F",
    "industry": ACCENT_RED,
    "agriculture": "6E8B1F",
    "lulucf": "007B6C",
    "finance": "A530B8",
    "fairness": ACCENT_ORANGE,
    "adaptation": "478EA5",
}
VERDICT_NOTE = {
    "CONFIRMED": "Reproduces from the primary source",
    "REVISION": "2–5% off the source — see the fact-check",
    "WRONG": "Does not reproduce — see the fact-check",
    "NO SOURCE YEAR": "The source has no figure for that year yet",
    "NOT CHECKABLE": "Published figure — source named, not re-derived",
}
VERDICT_COLOR = {
    "CONFIRMED": TEAL,
    "REVISION": ACCENT_ORANGE,
    "WRONG": ACCENT_RED,
    "NO SOURCE YEAR": MUTED,
    "NOT CHECKABLE": MUTED,
}
METHOD_LABEL = {
    "derived": "Derived from the live source",
    "reconstructed": "Reconstructed from the sheet's own conversion",
    "as-published": "Taken as published",
}


def read_indicator(ind):
    """Baseline / latest / change, exactly as the Method Hub module computes it."""
    data = sorted(ind["data"], key=lambda d: d["year"])
    pre = [d for d in data if not d.get("afterReport")]
    post = [d for d in data if d.get("afterReport")]
    baseline = pre[-1] if pre else None
    latest = data[-1] if data else None
    pct = None
    if baseline and latest and baseline["value"] != 0 and post:
        pct = (latest["value"] - baseline["value"]) / abs(baseline["value"]) * 100
    return {
        "ind": ind, "data": data, "pre": pre, "post": post,
        "baseline": baseline, "latest": latest, "pct": pct,
    }


def method_of(layout):
    """Classify a calc layout the way migration 079 built it."""
    if not layout:
        return None
    headers = [c["header"] for c in layout["columns"]]
    if any(h.startswith("Post-report figure as published") for h in headers):
        return "as-published"
    expr = layout["columns"][0].get("formula") or ""
    if expr.startswith("IF("):
        return "derived"
    return "reconstructed"


def factcheck_index(factcheck):
    """Per indicator: the worst verdict over its points, and how deep the check goes."""
    order = ["CONFIRMED", "NOT CHECKABLE", "NO SOURCE YEAR", "REVISION", "WRONG"]
    out = {}
    for r in factcheck.get("rows", []):
        cur = out.setdefault(r["id"], {"verdict": "CONFIRMED", "depth": r.get("depth"),
                                       "points": {}, "join": r.get("join")})
        if order.index(r["verdict"]) > order.index(cur["verdict"]):
            cur["verdict"] = r["verdict"]
        cur["depth"] = cur["depth"] or r.get("depth")
        cur["points"][r["year"]] = r
    return out


def build(data, calc, factcheck, out_path=None, wb=None, wire_targets=None):
    """
    Render the Indicator Check workbook.

    Standalone (default): `wb` is None, a fresh Workbook is created, the
    sheet's own "Read me" is written, and the result is saved to `out_path`.

    Embedded (combined workbook): pass an existing `wb` — its sheets are
    appended to (no "Read me" of its own; the caller supplies one merged
    "Read me" instead), and nothing is saved here. `out_path` is ignored.

    `wire_targets`, when given a list, has one dict appended per Overview
    latest-value cell, chapter-tab "In the 2024 report" / "Added since the
    report" cell, and Data (long) Value cell this build writes —
    {"sheet", "row", "col", "id", "year"} — for the combined workbook's
    derivation-wiring pass to rewrite into a cross-sheet formula wherever a
    Derivations block covers that (indicator, year). Left None (the default,
    and always so outside the combined workbook), nothing is recorded.
    """
    standalone = wb is None
    inds = [i for i in data["indicators"]]
    reads = {i["id"]: read_indicator(i) for i in inds}
    by_cat = defaultdict(list)
    for i in inds:
        by_cat[i["category"]].append(i)
    cats = [c for c in CATEGORY_ORDER if c in by_cat] + \
           [c for c in by_cat if c not in CATEGORY_ORDER]

    updated = [i for i in inds if reads[i["id"]]["post"]]
    fc = factcheck_index(factcheck)

    if standalone:
        wb = Workbook()

        # ── Read me ──────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Read me"
        sheet_setup(ws)
        set_widths(ws, [24, 22, 22, 22, 22, 20, 16, 16])
        r = title_block(
            ws,
            "Indicator Check — what has moved since the last report",
            f"Every progress indicator of the 2024 ESABCC report, with the latest data points added since "
            f"publication and the change against the report baseline. {len(updated)} of the {len(inds)} "
            f"indicators have gained data since {REPORT_YEAR}.",
            f"Background document for the Summer Prep · Policy Gap 2.0 Report module (Note 1) · prepared {PREPARED}.",
        )

        ws.cell(row=r, column=1, value="What is in this workbook").font = H2_FONT
        r += 1
        sheets = [
            ("Overview", f"All {len(inds)} indicators in one table — baseline, latest value, change — plus the two "
                         "overview figures: the largest moves since the report, and how many indicators have new "
                         "data per chapter."),
            ("One tab per chapter", "Emissions, Energy supply … Finance: the chapter's indicators as a table, then "
                                    "every indicator's full series as its own small table and line chart. The "
                                    "report's own data points and the points added since publication are drawn as "
                                    "two separate lines."),
            ("Data (long)", "Every data point of every indicator in one flat table — indicator, year, value, and "
                            "whether it is a report point or one added since. This is the sheet to pivot or filter."),
            ("Where the data comes from", "For every point added since the report: the calc inputs behind it, the "
                                          "formula that turns them into the plotted figure, the exact source "
                                          "query for each input, and the 27 July 2026 fact-check verdict — whether "
                                          "the figure still reproduces from that source today."),
        ]
        header_row(ws, r, ["Sheet", "What it holds", "", "", "", "", "", ""], height=20)
        r += 1
        for i, (name, desc) in enumerate(sheets):
            ws.cell(row=r, column=1, value=name).font = Font(name=FONT_SEMI, size=9, color=TEAL)
            ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
            c = ws.cell(row=r, column=2, value=desc)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = BODY_FILL if i % 2 == 0 else PatternFill("solid", fgColor=TEAL_PALE)
            ws.row_dimensions[r].height = 44
            r += 1
        r += 1

        ws.cell(row=r, column=1, value="How to read it").font = H2_FONT
        r += 1
        for line in [
            "• “Report baseline” is the latest figure the 2024 report itself carried. “Latest” is the most recent point "
            "in the database today. The change between them is arithmetic — no better/worse reading is applied, exactly "
            "as on the Method Hub page. Whether a move is good news depends on the indicator, and that judgement is yours.",
            "• A blank change column means the indicator has gained no data since the report.",
            "• Data vintages differ by indicator: an emissions figure for 2024 and an investment figure for 2025 are both "
            "“latest”, but they are not the same year. The year is always shown next to the value.",
            "• Every number here also lives in the Policy Gap 2.0 Project Workspace indicator database — this workbook "
            "shows nothing the workspace does not. If the two disagree, the workspace is the live copy.",
        ]:
            r = note_line(ws, r, line, width=8, font=BODY_FONT)
            ws.row_dimensions[r - 1].height = 30
        r += 1

        ws.cell(row=r, column=1, value="Source").font = H2_FONT
        r += 1
        for label, value in [
            ("Indicators", "The progress indicators of ESABCC (2024) “Towards EU climate neutrality: Progress, policy "
                           "gaps and opportunities” (O1–O3, E1–E6, I1–I7, T1–T6, B1–B6, A1–A7, L1–L8 and the finance / "
                           "innovation figures)."),
            ("Data source", "The Policy Gap 2.0 Project Workspace indicator database (src/data/esabcc-indicators.ts "
                            "and the calc grids in supabase/migrations/045, 052, 078, 079)."),
            ("Primary sources", "Eurostat, EEA, and the sector publications named per indicator — see the “Where the "
                                "data comes from” sheet for the exact query behind each post-report figure."),
            ("Vintage", f"Compiled {PREPARED}. Source publishers revise: re-check before quoting a figure externally."),
        ]:
            ws.cell(row=r, column=1, value=label).font = LABEL_FONT
            c = ws.cell(row=r, column=2, value=value)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 34
            r += 1

    # ── Overview ─────────────────────────────────────────────────────────────
    ov = wb.create_sheet("Overview")
    sheet_setup(ov, freeze="C6")
    headers = [
        "Code", "Indicator", "Chapter", "Unit", "Report baseline (year)", "Report baseline",
        "Latest (year)", "Latest value", "Change", "Change %", "New points since the report",
        "Source check (27 Jul 2026)", "Primary source",
    ]
    r = title_block(
        ov,
        "Every indicator, and what has moved",
        "Baseline = the last figure the 2024 report carried; latest = the most recent point in the database. "
        "The change is arithmetic only — no better/worse reading is applied.",
        width=13,
    )
    header_row(ov, r, headers, height=34)
    head_row = r
    set_widths(ov, [12, 46, 16, 14, 13, 13, 11, 12, 11, 11, 13, 30, 34])
    r += 1
    first = r
    for n, ind in enumerate(inds):
        rd = reads[ind["id"]]
        b, l = rd["baseline"], rd["latest"]
        row = r + n
        change = (l["value"] - b["value"]) if (b and l and rd["post"]) else None
        body_row(ov, row, [
            ind.get("code") or "",
            ind["name"],
            CATEGORY_LABEL.get(ind["category"], ind["category"]),
            ind["unit"],
            b["year"] if b else None,
            b["value"] if b else None,
            l["year"] if l else None,
            l["value"] if l else None,
            change,
            rd["pct"] / 100 if rd["pct"] is not None else None,
            len(rd["post"]) or None,
            (VERDICT_NOTE.get(fc[ind["id"]]["verdict"], "") if ind["id"] in fc else "—")
            if rd["post"] else "—",
            ind.get("source") or "",
        ], band=(n % 2 == 0), height=15)
        if wire_targets is not None and l:
            wire_targets.append({"sheet": ov.title, "row": row, "col": 8,
                                 "id": ind["id"], "year": l["year"]})
        ov.cell(row=row, column=1).font = Font(name=FONT_SEMI, size=9,
                                               color=CATEGORY_COLOR.get(ind["category"], TEAL))
        for col in (5, 7, 11):
            ov.cell(row=row, column=col).number_format = "0"
            ov.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        for col in (6, 8, 9):
            ov.cell(row=row, column=col).number_format = "#,##0.00"
            ov.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="center")
        c = ov.cell(row=row, column=10)
        c.number_format = "+0.0%;-0.0%;0.0%"
        c.alignment = Alignment(horizontal="right", vertical="center")
        if ind["id"] in fc and rd["post"]:
            ov.cell(row=row, column=12).font = Font(
                name=FONT_SEMI, size=8.5, color=VERDICT_COLOR[fc[ind["id"]]["verdict"]])
        else:
            ov.cell(row=row, column=12).font = SMALL_FONT
        src_cell = ov.cell(row=row, column=13)
        src_cell.font = SMALL_FONT
        link_cell(src_cell, detect_link(ind.get("source")))
    last = first + len(inds) - 1
    ov.auto_filter.ref = f"A{head_row}:M{last}"

    # Overview figure 1 — the largest moves since the report.
    movers = sorted((i for i in updated if reads[i["id"]]["pct"] is not None),
                    key=lambda i: abs(reads[i["id"]]["pct"]), reverse=True)[:18]
    fr = last + 3
    ov.cell(row=fr, column=1, value="Overview figures").font = H2_FONT
    fr += 1
    # The label goes in column B — the wide "Indicator" column — so the names
    # are readable in the sheet as well as in the figure. Painted only across
    # the 3 columns this mini-table uses: the figure is anchored at column E,
    # and a wider band would sit underneath it.
    header_row(ov, fr, ["", "Indicator (largest moves)", "Change % vs report baseline"], height=20)
    m_first = fr + 1
    for n, ind in enumerate(reversed(movers)):  # smallest first → largest on top of a bar chart
        rd = reads[ind["id"]]
        body_row(ov, m_first + n, ["", f"{ind.get('code')} · {ind['name']}", rd["pct"] / 100],
                 band=(n % 2 == 0), height=15)
        ov.cell(row=m_first + n, column=3).number_format = "+0.0%;-0.0%;0.0%"
        ov.cell(row=m_first + n, column=3).alignment = Alignment(horizontal="right", vertical="center")
    m_last = m_first + len(movers) - 1

    chart = BarChart()
    chart.type = "bar"
    chart.add_data(Reference(ov, min_col=3, min_row=fr, max_row=m_last), titles_from_data=True)
    chart.set_categories(Reference(ov, min_col=2, min_row=m_first, max_row=m_last))
    style_chart(chart, "The largest moves since the report baseline", height=11.5, width=20)
    text_categories(chart, ov, 2, m_first, m_last)
    chart.legend = None
    chart.x_axis.title = "Change vs the report's last figure (%)"
    chart.x_axis.numFmt = "0%"
    ser = chart.series[0]
    ser.data_points = [DataPoint(idx=i) for i in range(len(movers))]
    for i, ind in enumerate(reversed(movers)):
        pct = reads[ind["id"]]["pct"]
        ser.data_points[i].graphicalProperties.solidFill = TEAL if pct < 0 else ACCENT_ORANGE
        ser.data_points[i].graphicalProperties.line.noFill = True
    ov.add_chart(chart, f"E{fr}")

    # Overview figure 2 — coverage per chapter.
    cr = m_last + 3
    # Same reasoning as the movers header just above: painted only across the
    # 4 columns this mini-table uses, not the full 13-column width, so the
    # figure anchored at column E lands on empty cells.
    header_row(ov, cr, ["", "Chapter", "With new data since the report",
                        "Still at the report figure"], height=28)
    c_first = cr + 1
    for n, cat in enumerate(cats):
        up = sum(1 for i in by_cat[cat] if reads[i["id"]]["post"])
        body_row(ov, c_first + n, ["", CATEGORY_LABEL.get(cat, cat), up, len(by_cat[cat]) - up],
                 band=(n % 2 == 0), height=15)
        for col in (3, 4):
            ov.cell(row=c_first + n, column=col).alignment = Alignment(horizontal="center", vertical="center")
    c_last = c_first + len(cats) - 1

    cov = BarChart()
    cov.type = "col"
    cov.grouping = "stacked"
    cov.overlap = 100
    cov.add_data(Reference(ov, min_col=3, max_col=4, min_row=cr, max_row=c_last), titles_from_data=True)
    cov.set_categories(Reference(ov, min_col=2, min_row=c_first, max_row=c_last))
    style_chart(cov, "How much of each chapter has moved on since the report", height=9, width=20)
    text_categories(cov, ov, 2, c_first, c_last)
    cov.y_axis.title = "Indicators"
    for series, color in zip(cov.series, [TEAL, "D8E7E8"]):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.noFill = True
    cov.legend.position = "b"
    ov.add_chart(cov, f"E{cr}")

    note_line(ov, c_last + 2,
              "Notes: “change” is arithmetic — the workbook draws no conclusion about whether a move is progress. "
              "Latest years differ between indicators; the year is always shown beside the value.", width=12)
    note_line(ov, c_last + 3,
              "Source: ESABCC (2024) “Towards EU climate neutrality”, progress indicators, extended with the "
              "post-publication points held in the Policy Gap 2.0 indicator database.", width=13)

    # ── One tab per chapter ──────────────────────────────────────────────────
    for cat in cats:
        label = CATEGORY_LABEL.get(cat, cat)
        sh = wb.create_sheet(label[:31])
        color = CATEGORY_COLOR.get(cat, TEAL)
        sheet_setup(sh, tab_color=color)
        set_widths(sh, [11, 40, 13, 12, 11, 11, 11, 11, 26, 14, 14])
        r = title_block(
            sh, f"{label} — indicators",
            f"{len(by_cat[cat])} indicator(s). The table is the summary; below it, every indicator's full series "
            "with its own figure — the report's own points and the points added since publication as two lines.",
            width=11,
        )
        header_row(sh, r, ["Code", "Indicator", "Unit", "Baseline (yr)", "Baseline",
                           "Latest (yr)", "Latest", "Change %", "Primary source", "New points", "Provenance"],
                   height=30)
        r += 1
        for n, ind in enumerate(by_cat[cat]):
            rd = reads[ind["id"]]
            b, l = rd["baseline"], rd["latest"]
            layout = calc.get(ind["id"])
            body_row(sh, r + n, [
                ind.get("code") or "", ind["name"], ind["unit"],
                b["year"] if b else None, b["value"] if b else None,
                l["year"] if l else None, l["value"] if l else None,
                rd["pct"] / 100 if rd["pct"] is not None else None,
                ind.get("source") or "",
                len(rd["post"]) or None,
                METHOD_LABEL.get(method_of(layout), "—") if rd["post"] else "—",
            ], band=(n % 2 == 0), height=15)
            sh.cell(row=r + n, column=1).font = Font(name=FONT_SEMI, size=9, color=color)
            for col in (4, 6, 10):
                sh.cell(row=r + n, column=col).number_format = "0"
                sh.cell(row=r + n, column=col).alignment = Alignment(horizontal="center", vertical="center")
            for col in (5, 7):
                sh.cell(row=r + n, column=col).number_format = "#,##0.00"
                sh.cell(row=r + n, column=col).alignment = Alignment(horizontal="right", vertical="center")
            sh.cell(row=r + n, column=8).number_format = "+0.0%;-0.0%;0.0%"
            sh.cell(row=r + n, column=8).alignment = Alignment(horizontal="right", vertical="center")
            sh.cell(row=r + n, column=9).font = SMALL_FONT
            sh.cell(row=r + n, column=11).font = SMALL_FONT
        r += len(by_cat[cat]) + 2

        # Per-indicator series block + figure.
        for ind in by_cat[cat]:
            rd = reads[ind["id"]]
            data_rows = rd["data"]
            head = f"{ind.get('code')} · {ind['name']}  ({ind['unit']})"
            sh.cell(row=r, column=1, value=head).font = Font(name=FONT_SEMI, size=10, color=color)
            sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 1
            src = ind.get("source") or ""
            if ind.get("sourceUrl"):
                src = f"{src} — {ind['sourceUrl']}"
            c = sh.cell(row=r, column=1, value=f"Source: {src}")
            c.font = SMALL_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 1
            # Painted only across the 3 columns the table actually uses — the
            # figure sits at column E, and a header band running wider than
            # the table would float underneath it (a chart's anchor must land
            # on genuinely empty cells, not just visually-blank ones).
            header_row(sh, r, ["Year", "In the 2024 report", "Added since the report"], height=28)
            t_head = r
            r += 1
            t_first = r
            last_pre_idx = None
            for n, d in enumerate(data_rows):
                is_post = bool(d.get("afterReport"))
                if not is_post:
                    last_pre_idx = n
                body_row(sh, r + n, [
                    d["year"],
                    None if is_post else d["value"],
                    d["value"] if is_post else None,
                ], band=(n % 2 == 0), height=14)
                if wire_targets is not None:
                    wire_targets.append({"sheet": sh.title, "row": r + n,
                                         "col": 3 if is_post else 2,
                                         "id": ind["id"], "year": d["year"]})
                for col in (1, 2, 3):
                    sh.cell(row=r + n, column=col).alignment = Alignment(horizontal="right", vertical="center")
                sh.cell(row=r + n, column=1).number_format = "0"
                for col in (2, 3):
                    sh.cell(row=r + n, column=col).number_format = "#,##0.00"
            # Join the two lines: repeat the last report value in the post column.
            if last_pre_idx is not None and rd["post"]:
                sh.cell(row=t_first + last_pre_idx, column=3,
                        value=data_rows[last_pre_idx]["value"]).number_format = "#,##0.00"
                if wire_targets is not None:
                    wire_targets.append({"sheet": sh.title, "row": t_first + last_pre_idx,
                                         "col": 3, "id": ind["id"],
                                         "year": data_rows[last_pre_idx]["year"]})
            t_last = t_first + len(data_rows) - 1

            if len(data_rows) < 2:
                # One point is not a line: say so where the figure would be.
                c = sh.cell(row=t_head, column=5,
                            value="Only one data point — the report carries this indicator as a "
                                  "single figure, so there is no series to draw.")
                c.font = SMALL_FONT
                c.alignment = Alignment(vertical="top", wrap_text=True)
                sh.merge_cells(start_row=t_head, start_column=5, end_row=t_head + 1, end_column=9)
                r = max(t_last + 2, t_head + 3)
                continue

            ch = LineChart()
            ch.add_data(Reference(sh, min_col=2, max_col=3, min_row=t_head, max_row=t_last),
                        titles_from_data=True)
            ch.set_categories(Reference(sh, min_col=1, min_row=t_first, max_row=t_last))
            style_chart(ch, f"{ind.get('code')} — {ind['name']}", height=6.6, width=13.5,
                       y_number_format="#,##0")
            ch.y_axis.title = ind["unit"]
            ch.series[0].graphicalProperties.line.solidFill = color
            ch.series[0].graphicalProperties.line.width = 22000
            ch.series[0].smooth = False
            ch.series[0].marker = Marker(symbol="none")
            post_ser = ch.series[1]
            post_ser.graphicalProperties.line.solidFill = ACCENT_ORANGE
            post_ser.graphicalProperties.line.width = 22000
            post_ser.graphicalProperties.line.dashStyle = "dash"
            post_ser.smooth = False
            post_ser.marker = Marker(symbol="circle", size=5)
            ch.legend.position = "b"
            sh.add_chart(ch, f"E{t_head}")

            r = max(t_last + 2, t_head + 14)

    # ── Data (long) ──────────────────────────────────────────────────────────
    lg = wb.create_sheet("Data (long)")
    sheet_setup(lg, freeze="A2")
    header_row(lg, 1, ["Code", "Indicator", "Chapter", "Unit", "Year", "Value",
                       "Point origin", "Primary source", "Indicator id"], height=22)
    set_widths(lg, [11, 42, 16, 14, 8, 13, 20, 34, 26])
    row = 2
    for ind in inds:
        for d in sorted(ind["data"], key=lambda x: x["year"]):
            post = bool(d.get("afterReport"))
            body_row(lg, row, [
                ind.get("code") or "", ind["name"],
                CATEGORY_LABEL.get(ind["category"], ind["category"]), ind["unit"],
                d["year"], d["value"],
                "Added since the report" if post else "In the 2024 report",
                ind.get("source") or "", ind["id"],
            ], band=(row % 2 == 0), height=14)
            if wire_targets is not None:
                wire_targets.append({"sheet": lg.title, "row": row, "col": 6,
                                     "id": ind["id"], "year": d["year"]})
            lg.cell(row=row, column=5).number_format = "0"
            lg.cell(row=row, column=6).number_format = "#,##0.000"
            lg.cell(row=row, column=7).font = Font(name=FONT_SEMI, size=9,
                                                   color=ACCENT_ORANGE if post else TEAL)
            lg.cell(row=row, column=9).font = SMALL_FONT
            row += 1
    lg.auto_filter.ref = f"A1:I{row - 1}"

    # ── Where the data comes from ────────────────────────────────────────────
    pv = wb.create_sheet("Where the data comes from")
    sheet_setup(pv, freeze="B6")
    set_widths(pv, [11, 34, 8, 13, 22, 18, 30, 15, 46, 26])
    r = title_block(
        pv,
        "Where each post-report number comes from",
        "For every data point added since the 2024 report: the inputs the figure is built from, the value of each "
        "input, and the exact source query behind it. This is the same calc space the Method Hub shows behind "
        "“Edit data / calc”.",
        width=8,
    )
    header_row(pv, r, ["Code", "Indicator", "Year", "Value", "How this year was obtained",
                       "Checked 27 Jul 2026", "Input", "Input value",
                       "Where that input comes from", "Exact dataset link (DOI where available)"], height=32)
    r += 1
    n = 0
    for ind in inds:
        rd = reads[ind["id"]]
        if not rd["post"]:
            continue
        layout = calc.get(ind["id"])
        method = METHOD_LABEL.get(method_of(layout), "Taken as published")
        for d in rd["post"]:
            year = d["year"]
            row_cells = None
            if layout:
                row_cells = next((rw for rw in layout["rows"] if rw["year"] == year), None)
            inputs = []
            if row_cells:
                for ci, col in enumerate(layout["columns"]):
                    if ci == 0:
                        continue
                    v = row_cells["cells"][ci] if ci < len(row_cells["cells"]) else None
                    if isinstance(v, dict):
                        v = v.get("v")
                    if v is None:
                        continue
                    inputs.append((col["header"], v, col.get("source") or ""))
            if not inputs:
                inputs = [("(no calc inputs recorded)", None, ind.get("source") or "")]
            for k, (header, value, source) in enumerate(inputs):
                point = fc.get(ind["id"], {}).get("points", {}).get(year)
                verdict = point["verdict"] if point else "NOT CHECKABLE"
                body_row(pv, r, [
                    ind.get("code") if k == 0 else "",
                    ind["name"] if k == 0 else "",
                    year if k == 0 else None,
                    d["value"] if k == 0 else None,
                    method if k == 0 else "",
                    (f"{verdict} — {VERDICT_NOTE[verdict]}" if k == 0 else ""),
                    header, value, source,
                ], band=(n % 2 == 0), height=26)
                pv.cell(row=r, column=3).number_format = "0"
                pv.cell(row=r, column=4).number_format = "#,##0.000"
                pv.cell(row=r, column=8).number_format = "#,##0.000"
                pv.cell(row=r, column=8).alignment = Alignment(horizontal="right", vertical="top")
                pv.cell(row=r, column=1).font = Font(name=FONT_SEMI, size=9, color=TEAL)
                if k == 0:
                    pv.cell(row=r, column=6).font = Font(name=FONT_SEMI, size=8.5,
                                                         color=VERDICT_COLOR[verdict])
                pv.cell(row=r, column=9).font = SMALL_FONT
                pv.cell(row=r, column=7).font = Font(name=FONT_LIGHT, size=9, color=INK)
                link_cell(pv.cell(row=r, column=10), detect_link(source))
                r += 1
            n += 1
        # the formula that turns those inputs into the plotted value
        if layout and layout["columns"][0].get("formula"):
            c = pv.cell(row=r, column=7, value=f"Formula:  Value = {layout['columns'][0]['formula']}")
            c.font = Font(name=FONT_LIGHT, size=8, color=MUTED)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            pv.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
            pv.row_dimensions[r].height = 22
            r += 1
    pv.auto_filter.ref = f"A5:J{r - 1}"
    note_line(pv, r + 1,
              "Notes: “Derived from the live source” means the figure is recomputed from the publisher's own data, "
              "with the query in the last column. “Reconstructed” means the sheet's own conversion pins the input "
              "exactly from the published figure. “Taken as published” means the figure is the publication's own — "
              "no derivation is claimed. The verdict column is the 27 July 2026 fact-check: whether the stored "
              "figure still reproduces from that source today. Two series (L1, L7) reproduce their source but sit "
              "on a different level than the report years — see "
              "docs-internal/indicator-postreport-factcheck-2026-07-27.md.", width=9)

    if standalone:
        # No cached results are written for the COUNTIFS formulas, so ask the
        # spreadsheet app to calculate the whole book the moment it opens.
        wb.calculation.fullCalcOnLoad = True
        wb.properties.title = "ESABCC Indicator Check — what has moved since the 2024 report"
        wb.properties.subject = "Summer Prep · Policy Gap 2.0 Report — background document"
        wb.properties.creator = "ESABCC Method Hub"
        wb.save(out_path)
        print(f"wrote {out_path}: {len(inds)} indicators, {len(updated)} with new data, "
              f"{len(cats)} chapter tabs")
    return wb


if __name__ == "__main__":
    build(json.load(open(sys.argv[1], encoding="utf8")),
          json.load(open(sys.argv[2], encoding="utf8")),
          json.load(open(sys.argv[3], encoding="utf8")),
          sys.argv[4])
