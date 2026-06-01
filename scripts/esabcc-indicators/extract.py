"""Extract indicator data from the ESABCC report Excel underlying data file.

Reads each figure sheet from the workbook, looks for the canonical
``Label / Unit / <years> / Source`` header row produced by the report
authors, and dumps each data series to JSON. Large pages (>200 used
rows) are truncated since they hold disaggregated micro-data, not
plotted series, and would otherwise blow the read_only iterator.
"""
import openpyxl
import json
import sys

INFILE = 'ESABCC_report_Towards EU climate neutrality_underlying data.xlsx'
MAX_ROWS = 200
MAX_HEADER_SEARCH = 60

wb = openpyxl.load_workbook(INFILE, data_only=True, read_only=True)

# Build figure -> indicator title map from Index sheet
index = {}
ws = wb['Index']
for r in ws.iter_rows(values_only=True):
    if not r or not r[0] or r[0] == 'Worksheet':
        continue
    sheet = str(r[0]).strip()
    title = str(r[1]).strip() if r[1] else ''
    index[sheet] = title

result = {}

for sn, title in index.items():
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i >= MAX_ROWS:
            break
        rows.append(list(r))

    header_idx = None
    for i, r in enumerate(rows[:MAX_HEADER_SEARCH]):
        if not r:
            continue
        first_str = str(r[0]).strip() if r[0] else ''
        if first_str.lower() == 'label':
            header_idx = i
            break

    if header_idx is None:
        # Fallback: row with many year-like ints
        for i, r in enumerate(rows[:MAX_HEADER_SEARCH]):
            if not r:
                continue
            years = sum(
                1 for c in r if isinstance(c, (int, float)) and 1900 < c < 2100
            )
            if years >= 5:
                header_idx = i
                break

    if header_idx is None:
        result[sn] = {'title': title, 'series': [], 'note': 'no header'}
        continue

    header = rows[header_idx]
    year_cols = [
        (j, int(c))
        for j, c in enumerate(header)
        if isinstance(c, (int, float)) and 1900 < c < 2100
    ]
    source_col = None
    for j, c in enumerate(header):
        if isinstance(c, str) and c.strip().lower() == 'source':
            source_col = j
            break

    series = []
    for i in range(header_idx + 1, len(rows)):
        r = rows[i]
        if not r or all(c is None for c in r):
            continue
        label = r[0] if len(r) > 0 else None
        unit = r[1] if len(r) > 1 else None
        if not isinstance(label, str) or not label.strip():
            continue
        data = []
        for j, y in year_cols:
            if j < len(r) and isinstance(r[j], (int, float)):
                data.append({'year': y, 'value': float(r[j])})
        src = (
            r[source_col]
            if source_col is not None and source_col < len(r)
            else None
        )
        series.append(
            {
                'label': label.strip(),
                'unit': str(unit).strip() if unit else None,
                'source': str(src).strip() if src else None,
                'data': data,
            }
        )
    result[sn] = {'title': title, 'series': series}

json.dump(result, sys.stdout, indent=2, default=str)
