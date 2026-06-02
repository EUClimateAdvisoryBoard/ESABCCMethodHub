"""Reverse-engineer ESABCC indicator derivations from the chapter workbooks
into the Indicator Database calc-space layout (pw_indicator_sheets.layout).

A *recipe* describes, for one indicator, how its published series is derived
from raw inputs found in a chapter sheet. We build a year-indexed grid:

    Year | Value | helper_1 | helper_2 | ...

where `Value` carries a free-form formula (e.g. "[Raw (Mtoe)] * [Mtoe->TWh]")
that documents the operation, and the helper columns hold the raw inputs.
This is exactly the shape src/lib/project-workspace/indicator-sheet.ts expects,
so the derivation recomputes live in the app and re-exports as a live Excel
formula. Drop in next year's raw data and the indicator reproduces itself.
"""
import openpyxl, re, json

CH_DIR = 'scratch/ch'

_wb_cache = {}
def load(path, data_only):
    key = (path, data_only)
    if key not in _wb_cache:
        _wb_cache[key] = openpyxl.load_workbook(path, data_only=data_only, read_only=True)
    return _wb_cache[key]

def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26); s = chr(65 + r) + s
    return s

def grid(path, sheet):
    """Return (values, formulas) 2D dicts keyed [row][col] (1-based)."""
    wv = load(path, True)[sheet]
    wf = load(path, False)[sheet]
    V, F = {}, {}
    for row in wv.iter_rows():
        for c in row:
            if c.value is not None: V[(c.row, c.column)] = c.value
    for row in wf.iter_rows():
        for c in row:
            if c.value is not None: F[(c.row, c.column)] = c.value
    return V, F

# ── expr evaluator (mirrors the subset of formula.ts we emit) ────────────────
def eval_expr(expr, headers, values):
    """Evaluate a Value formula for one row. `values` maps header->number.
    Supports [Header] / bare refs, + - * / and parentheses."""
    def repl(m):
        name = m.group(1)
        v = values.get(name)
        return f'({v})' if v is not None else 'None'
    s = re.sub(r'\[([^\]]+)\]', repl, expr)
    # bare single-word refs
    for h in sorted(headers, key=len, reverse=True):
        if re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', h):
            s = re.sub(rf'(?<![\[\w]){re.escape(h)}(?![\w\]])',
                       f'({values[h]})' if values.get(h) is not None else 'None', s)
    try:
        return eval(s, {"__builtins__": {}}, {})
    except Exception:
        return None

# ── layout assembly ──────────────────────────────────────────────────────────
def build_layout(value_header, value_source, value_formula, helpers, years, raw):
    """helpers: list of {header, source}. raw: (header, year)->number.
    Returns (layout_dict, computed_by_year)."""
    columns = [{"header": value_header, "source": value_source,
                "formula": {"expr": value_formula}}]
    for h in helpers:
        columns.append({"header": h["header"], "source": h.get("source")})
    headers = [value_header] + [h["header"] for h in helpers]
    rows, computed = [], {}
    for y in years:
        hv = {h["header"]: raw(h["header"], y) for h in helpers}
        val = eval_expr(value_formula, headers, hv)
        computed[y] = val
        cells = [val] + [hv[h["header"]] for h in helpers]
        rows.append({"year": y, "cells": cells})
    # strip null sources
    for c in columns:
        if c.get("source") in (None, ""): c.pop("source", None)
    return {"columns": columns, "rows": rows}, computed

def verify(name, computed, expected, tol=0.5):
    """Compare computed[year] against expected[year] (overview/cached)."""
    msgs, ok = [], True
    common = sorted(set(computed) & set(expected))
    if not common:
        return False, [f"  {name}: NO overlapping years to verify"]
    worst = 0.0
    for y in common:
        cv, ev = computed[y], expected[y]
        if cv is None:
            ok = False; msgs.append(f"  {name} {y}: computed None (expected {ev})"); continue
        d = abs(cv - ev); worst = max(worst, d)
        if d > tol:
            ok = False
            msgs.append(f"  {name} {y}: computed {cv:.3f} vs expected {ev:.3f} (Δ{d:.3f})")
    head = f"  {name}: {'OK' if ok else 'MISMATCH'} ({len(common)} yrs, worst Δ{worst:.4f})"
    return ok, [head] + msgs
