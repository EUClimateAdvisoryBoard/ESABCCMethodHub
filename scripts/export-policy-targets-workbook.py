#!/usr/bin/env python3
"""Export the M·36 Policy Targets Register to the reviewer workbook.

    python3 scripts/export-policy-targets-workbook.py [out.xlsx]

Sheets
    Policy targets   one row per target of the corrected register: the original
                     twelve columns, plus the eight sector/system flags, the
                     mitigation/adaptation argument and the duplicate column.
                     Targets that fall in none of the eight sectors are kept —
                     the reviewer asked for them to stay visible.
    Removed rows     every row the July-2026 human-review pass removed, with the
                     rule and the reason, so each correction can be checked.
    Corrections log  the correction rules, what they mean and how many rows each
                     removed.
    Sector summary   descriptive statistics per sector/system.
    <Sector> sheets  the policies contributing to each sector, with counts and
                     an example target.

Input is the generated dataset (src/data/policy-targets.generated.ts), so run
`npm run build:policy-targets` first if the overrides or the classifier changed.
Requires openpyxl.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATASET = ROOT / 'src/data/policy-targets.generated.ts'
OVERRIDES = ROOT / 'scripts/policy-targets-overrides.json'
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / 'public/data/eu-policy-targets-corrected.xlsx'

SECTORS = [
    ('energy', 'Energy'),
    ('buildings', 'Buildings and built environment'),
    ('agrifood', 'Agri-food'),
    ('transport', 'Transport and mobility'),
    ('industry', 'Industry'),
    ('ecosystems', 'Land and marine ecosystems'),
    ('water', 'Water'),
    ('health', 'Health'),
]
SHEET_NAME = {  # Excel caps sheet names at 31 characters
    'buildings': 'Buildings & built environment',
    'ecosystems': 'Land & marine ecosystems',
}
CLIMATE_LABEL = {
    'mitigation': 'Mitigation', 'adaptation': 'Adaptation',
    'both': 'Mitigation + Adaptation', 'none': 'Not climate-specific',
}
RULES = [
    ('NT-1', 'Plan or report content requirement',
     'An item on the list of what a Member State plan, strategy or report must contain is not itself a '
     'target. Kept where the item carries its own date or quantity.'),
    ('NT-2', 'Eligibility or funding scope',
     'A list of measures or costs a fund may cover, or a condition for receiving money, is not a target.'),
    ('NT-3', 'Assessment, award or selection criteria',
     'Criteria for assessing a plan or selecting a project, and descriptions of what an eligible project '
     'looks like, are not targets.'),
    ('NT-4', 'Definition',
     'Definitions, and the chapeaux that introduce them, are not targets.'),
    ('NT-5', 'Pointer to another instrument’s target',
     'A requirement to be “in line with” a target set elsewhere is not a target of this act; the target is '
     'registered under the act that sets it.'),
    ('NT-6', 'Context or illustrative material',
     'Soft-law text summarising commitments made elsewhere, illustrative statistics and annex headings are '
     'not targets.'),
    ('DUP', 'Duplicate row',
     'The same sentence was already registered under a more precise provision reference.'),
    ('FIX', 'Superseded by a corrected extraction',
     'The quote was truncated — cut off at a colon, or opening with a pronoun that had no antecedent — and '
     'was re-extracted in full from the source.'),
    ('FIX-COLON', 'Chapeau completed',
     'The quote stopped at the colon introducing a list, and the enumerated points were not separate rows; '
     're-extracted with the points.'),
    ('FIX-ANAPHORA', 'Antecedent restored',
     'The quote opened on “It”, “This”, “Those” or similar with no antecedent; re-extracted with the '
     'preceding sentence.'),
    ('FIX-TRUNCATION', 'Sentence completed',
     'The quote was sliced mid-sentence — its subject or its continuation sat outside the slice; '
     're-extracted to sentence bounds.'),
]
PASS_LABEL = {
    '[human review 2026-07]': 'Pass 1 — reviewer mark-up generalised',
    '[agent review 2026-07 round 2]': 'Pass 2 — full agent audit',
    '[human review 2026-08 v3]': 'Pass 3 — v3 reviewer mark-up (Aug 2026)',
    '[re-extraction 2026-08]': 'Pass 4 — re-extraction from replaced consolidated texts (Aug 2026)',
    '[carried to consolidated text 2026-08]': 'Pass 4 — re-extraction from replaced consolidated texts (Aug 2026)',
}

HEADER_FILL = PatternFill('solid', fgColor='FFE8EEF4')
HEADER_FONT = Font(name='Arial', bold=True, size=10)
BODY_FONT = Font(name='Arial', size=10)


def load_targets() -> list[dict]:
    src = DATASET.read_text(encoding='utf8')
    marker = 'RAW_POLICY_TARGETS: RawPolicyTarget[] = '
    body = src[src.index(marker) + len(marker):]
    return json.loads(body[:body.rindex(']') + 1])


def cap(s: str) -> str:
    return s[:1].upper() + s[1:] if s else s


def write_sheet(ws, rows, widths, freeze='A2'):
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    if freeze:
        ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    targets = load_targets()
    rows = [t for t in targets if t['relevant']]
    overrides = json.loads(OVERRIDES.read_text(encoding='utf8'))
    review_drops = [o for o in overrides
                    if o.get('drop') and any(tag in (o.get('reason') or '') for tag in PASS_LABEL)]

    wb = Workbook()

    # ── Policy targets ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Policy targets'
    # Columns 12 ("Human confirmed") and the two mark-up columns at the far right
    # are left empty on purpose: they are where the reviewer records the next
    # round of confirmations and objections, as in the workbook they marked up.
    header = [
        '1 · Name of policy', '2 · Type of policy', 'Document updated (consolidated version)', '3 · Policy area',
        '#', 'First order target (1) Second order target (2)', '4 · Target text (verbatim)',
        'Provision', '5 · Target label', '6 · Obligation', '7 · Type of target', '8 · Timeline',
        '9 · Indicators', '10 · Climate-relevance', '11 · Source (EUR-Lex)', 'Relevant (transition lens)',
        'Revise target', 'Revise reason',
        '12 · Human confirmed',
    ] + [f'{13 + i} · {label}' for i, (_, label) in enumerate(SECTORS)] + [
        '21 · Mitigation / adaptation argument',
        '22 · Duplicate / similar target in another policy',
        'Not correct', 'reason',
    ]
    body = [[
        t['policy_name'], cap(t['document_type']), t.get('doc_replaced') or '', t['policy_area'], t['target_number'],
        t['target_order'] if t.get('target_order') in (1, 2) else '', t['target_text'],
        t['article'], cap(t['target_label']), cap(t['obligation']), cap(t['target_type']),
        t['timeline'] or 'Unspecified', '; '.join(t.get('indicators') or []),
        CLIMATE_LABEL[t['climate_relevance']], t['eurlex_url'], 'Relevant',
        1 if t.get('revise_flag') else '', t.get('revise_reason') or '', '',
    ] + ['Yes' if key in t['sectors'] else '' for key, _ in SECTORS] + [
        t['climate_argument'], t['duplicate_of'], '', '',
    ] for t in rows]
    write_sheet(ws, [header] + body,
                [42, 15, 30, 16, 5, 16, 70, 26, 13, 12, 14, 20, 28, 20, 40, 16, 12, 40, 16] + [13] * 8 + [62, 46, 12, 46])

    # ── Removed rows ────────────────────────────────────────────────────────
    removed = [['Row id', 'Pass', 'Rule', 'Reason for removal']]
    for o in review_drops:
        reason = o.get('reason') or ''
        rule, _, rest = reason.partition(':')
        which = next(label for tag, label in PASS_LABEL.items() if tag in reason)
        for tag in PASS_LABEL:
            rest = rest.replace(tag, '')
        removed.append([o['id'], which, rule.strip(), rest.strip()])
    write_sheet(wb.create_sheet('Removed rows'), removed, [16, 30, 15, 110])

    # ── Corrections log ─────────────────────────────────────────────────────
    counts = Counter((o.get('reason') or '').partition(':')[0].strip() for o in review_drops)
    log = [['Rule', 'Name', 'What it means', 'Rows affected']]
    log += [[k, n, d, counts.get(k, 0)] for k, n, d in RULES]
    log.append(['', 'Total rows removed', '', sum(counts.values())])
    write_sheet(wb.create_sheet('Corrections log'), log, [8, 38, 92, 14])

    # ── Sector summary + per-sector sheets ──────────────────────────────────
    def pct(n, d):
        return f'{round(n / d * 1000) / 10}%' if d else '—'

    summary = [[
        'Sector / system', 'Lens', 'Targets', '% of register', 'Policies', 'Mandatory', 'Quantitative',
        'With a timeline', 'Mitigation', 'Adaptation', 'Mitigation + Adaptation', 'Not climate-specific',
        'Also in another sector', 'Top policies (targets)',
    ]]
    sector_sheets = []
    for key, label in SECTORS:
        rs = [t for t in rows if key in t['sectors']]
        by_policy = Counter(t['policy_name'] for t in rs)
        summary.append([
            label,
            'Adaptation-critical' if key in ('water', 'health') else 'Mitigation + adaptation',
            len(rs), pct(len(rs), len(rows)), len(by_policy),
            sum(t['obligation'] == 'mandatory' for t in rs),
            sum(t['target_type'] == 'quantitative' for t in rs),
            sum(bool(t['timeline']) for t in rs),
            sum(t['climate_relevance'] == 'mitigation' for t in rs),
            sum(t['climate_relevance'] == 'adaptation' for t in rs),
            sum(t['climate_relevance'] == 'both' for t in rs),
            sum(t['climate_relevance'] == 'none' for t in rs),
            sum(len(t['sectors']) > 1 for t in rs),
            '; '.join(f'{n} ({c})' for n, c in by_policy.most_common(3)),
        ])
        sheet_rows = [['Policy', 'Type', 'Targets in this sector', 'Mandatory', 'Quantitative', 'Mitigation',
                       'Adaptation', 'Both', 'Not climate-specific', 'Earliest / example target']]
        for name, _ in by_policy.most_common():
            ps = [t for t in rs if t['policy_name'] == name]
            example = next((t for t in ps if t['target_type'] == 'quantitative'), ps[0])
            sheet_rows.append([
                name, cap(ps[0]['document_type']), len(ps),
                sum(t['obligation'] == 'mandatory' for t in ps),
                sum(t['target_type'] == 'quantitative' for t in ps),
                sum(t['climate_relevance'] == 'mitigation' for t in ps),
                sum(t['climate_relevance'] == 'adaptation' for t in ps),
                sum(t['climate_relevance'] == 'both' for t in ps),
                sum(t['climate_relevance'] == 'none' for t in ps),
                example['target_text'][:300],
            ])
        sector_sheets.append((SHEET_NAME.get(key, label), sheet_rows))

    no_sector = [t for t in rows if not t['sectors']]
    summary.append([])
    summary.append([
        'No sector / system (cross-cutting, institutional or financial)', '—', len(no_sector),
        pct(len(no_sector), len(rows)), len({t['policy_name'] for t in no_sector}),
        sum(t['obligation'] == 'mandatory' for t in no_sector),
        sum(t['target_type'] == 'quantitative' for t in no_sector),
        sum(bool(t['timeline']) for t in no_sector),
        sum(t['climate_relevance'] == 'mitigation' for t in no_sector),
        sum(t['climate_relevance'] == 'adaptation' for t in no_sector),
        sum(t['climate_relevance'] == 'both' for t in no_sector),
        sum(t['climate_relevance'] == 'none' for t in no_sector),
        0, 'Kept in the register for review, as requested',
    ])
    write_sheet(wb.create_sheet('Sector summary'), summary,
                [46, 22, 9, 13, 9, 11, 12, 14, 11, 11, 21, 19, 19, 60])
    for name, sheet_rows in sector_sheets:
        write_sheet(wb.create_sheet(name[:31]), sheet_rows, [46, 14, 12, 11, 12, 11, 11, 8, 16, 70])

    # ── About ───────────────────────────────────────────────────────────────
    acts = len({t['policy_id'] for t in rows})
    about = [
        ['EU policy targets register — corrected edition', ''],
        ['', ''],
        ['Built from', 'src/data/policy-targets.generated.ts (npm run build:policy-targets)'],
        ['Exported by', 'scripts/export-policy-targets-workbook.py'],
        ['Corrections documented in', 'docs-internal/policy-targets-human-review-2026-07.md'],
        ['', ''],
        ['Sheet', 'What it holds'],
        ['Policy targets', f'{len(rows)} targets across {acts} acts, after the July-2026 human-review corrections.'],
        ['Removed rows', f'{len(review_drops)} rows removed in that pass, each with the rule and the reason.'],
        ['Corrections log', 'The eight correction rules, what they mean, and how many rows each removed.'],
        ['Sector summary', 'Descriptive statistics per sector/system, plus the cross-cutting remainder.'],
        ['Sector sheets', 'The policies contributing targets to each sector, with counts and an example target.'],
        ['', ''],
        ['Columns 13-20', 'The eight sectors/systems. The first six matter for both mitigation and adaptation; '
                          'Water and Health are the two that matter above all for adaptation. A target can carry '
                          'several. Targets in none of the eight are kept for review.'],
        ['Column 21', 'The mechanism-based argument for the mitigation/adaptation call: cuts GHGs, raises removals '
                      'or enables either (mitigation); reduces vulnerability or exposure, raises adaptive capacity '
                      'or seizes an opportunity (adaptation).'],
        ['Column 22', 'The closest target text in another policy — “Duplicate wording” at ≥0.45 token similarity, '
                      '“Similar target” at ≥0.25 — with a note when the other act is named in the text.'],
        ['Column 12 and the last two columns', 'Left empty for the next review round: mark “12 · Human confirmed”, '
                                               'and use “Not correct” / “reason” to flag a row exactly as in the '
                                               'previous mark-up — the reasons written there become the rules of the '
                                               'next correction pass.'],
        ['Document updated (consolidated version)', 'Set when the act’s source document was replaced or '
                                                     'consolidated since extraction (scripts/policy-targets-replaced.json); '
                                                     'notes what changed.'],
        ['First order target (1) Second order target (2)', 'The v3 reviewer’s first/second-order call '
                                                            '(scripts/policy-targets-review-2026-08.json): 1 for the '
                                                            'headline change the act exists to achieve, 2 for a target '
                                                            'that depends on or complements one. Blank when unlabelled.'],
        ['Revise target / Revise reason', 'The v3 reviewer’s "likely not a target" flag and the reason '
                                          '(scripts/policy-targets-review-2026-08.json, `revise`).'],
    ]
    write_sheet(wb.create_sheet('About'), about, [30, 110], freeze=None)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f'✓ {len(rows)} targets across {acts} acts → {OUT.relative_to(ROOT) if OUT.is_relative_to(ROOT) else OUT}')
    print(f'  removed rows: {len(review_drops)}; sector sheets: {len(sector_sheets)}')


if __name__ == '__main__':
    main()
